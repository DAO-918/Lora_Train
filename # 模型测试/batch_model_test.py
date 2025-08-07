import os
import time
import io
import logging
import sys
from openpyxl import load_workbook
from PIL import Image as PILImage
# 将项目根目录添加到系统路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils import ComfyApiWrapper, ComfyWorkflowWrapper, ComfyWebSocketClient

# 配置日志
logging.basicConfig(stream=sys.stdout, level=logging.INFO)
logger = logging.getLogger(__name__)

class BatchModelTester:
    def __init__(self, model_info_path, comfy_host="127.0.0.1:8191"):
        """
        初始化批量测试模型类
        
        Args:
            model_info_path: 模型信息表格路径
            comfy_host: ComfyUI服务器地址
        """
        self.model_info_path = model_info_path
        self.model_info_wb = load_workbook(model_info_path, data_only=True)
        self.comfy_host = comfy_host
        self.ws_client = None
    
    def parse_value(self, value):
        """
        解析参数值，如果可以解析为数组则返回数组
        
        Args:
            value: 要解析的值
            
        Returns:
            解析后的值，如果可以解析为数组则返回数组，否则返回原值
        """
        if isinstance(value, str) and value.strip().startswith("[") and value.strip().endswith("]"):
            try:
                # 尝试解析为JSON数组
                import json
                return json.loads(value)
            except Exception as e:
                logger.warning(f"解析数组值失败: {e}, 使用原始值")
        return value
        
    def load_model_info(self):
        """
        加载模型信息表格中的数据
        """
        # 加载Stable-diffusion工作簿数据
        if "Stable-diffusion" in self.model_info_wb.sheetnames:
            sd_sheet = self.model_info_wb["Stable-diffusion"]
            headers = [cell.value for cell in sd_sheet[1]]
            self.sd_models = []
            for row in sd_sheet.iter_rows(min_row=2, values_only=True):
                model_data = {headers[i]: row[i] for i in range(len(headers))}
                if model_data.get("文件名"):
                    self.sd_models.append(model_data)
        else:
            self.sd_models = []
            logger.warning("模型信息表格中没有Stable-diffusion工作簿")
        
        # 加载Lora工作簿数据
        if "Lora" in self.model_info_wb.sheetnames:
            lora_sheet = self.model_info_wb["Lora"]
            headers = [cell.value for cell in lora_sheet[1]]
            self.lora_models = []
            for row in lora_sheet.iter_rows(min_row=2, values_only=True):
                model_data = {headers[i]: row[i] for i in range(len(headers))}
                if model_data.get("文件名"):
                    self.lora_models.append(model_data)
        else:
            self.lora_models = []
            logger.warning("模型信息表格中没有Lora工作簿")
    
    def process_test_file(self, test_file_path):
        """
        处理测试文件
        
        Args:
            test_file_path: 测试文件路径
        """
        # 加载测试文件
        workbook = load_workbook(test_file_path, data_only=True)
        
        # 检查必要的工作簿是否存在
        required_sheets = ["参数"]
        for sheet_name in required_sheets:
            if sheet_name not in workbook.sheetnames:
                logger.error(f"测试文件中缺少{sheet_name}工作簿")
                return
                
        # 检查是否存在底模工作簿
        has_base_models = "底模" in workbook.sheetnames
        
        # 加载参数工作簿
        params_sheet = workbook["参数"]
        params = {}
        for row in params_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 名称不为空
                # 解析参数值，如果可以解析为数组则返回数组
                value = self.parse_value(row[1])
                params[row[0]] = {
                    "值": value,
                    "节点名称": row[2],
                    "节点属性": row[3]
                }
        
        # 获取关键参数
        workflow_file = params.get("workflow", {}).get("值")
        default_model = params.get("默认底模", {}).get("值")
        prompt_id = params.get("提示词编号", {}).get("值")
        save_path = params.get("保存图片路径", {}).get("值")
        
        if not all([workflow_file, default_model, prompt_id, save_path]):
            logger.error("参数工作簿中缺少必要参数")
            return
        
        # 确保保存路径存在
        prompt_save_path = os.path.join(save_path, prompt_id)
        os.makedirs(prompt_save_path, exist_ok=True)
        
        # 加载工作流
        try:
            workflow = ComfyWorkflowWrapper(f'workflow/{workflow_file}')
        except Exception as e:
            logger.error(f"加载工作流失败: {e}")
            return
        
        # 处理底模工作簿(如果存在)
        if has_base_models:
            self.process_base_models(workbook, workflow, params, prompt_save_path)
        
        # 处理Lora工作簿(如果存在)
        lora_sheets = [sheet for sheet in workbook.sheetnames if sheet.startswith("Lora-")]
        if lora_sheets:
            self.process_lora_models(workbook, workflow, params, prompt_save_path, lora_sheets)
        
        # 处理测试工作簿(如果存在且有底模工作簿)
        if has_base_models and "测试" in workbook.sheetnames:
            self.process_test_combinations(workbook, workflow, params, prompt_save_path)
        
        # 保存工作簿
        workbook.save(test_file_path)
        logger.info(f"测试完成，结果已保存到 {test_file_path}")
    
    def process_base_models(self, workbook, workflow, params, save_path):
        """
        处理底模工作簿
        """
        base_sheet = workbook["底模"]
        
        # 获取列索引
        headers = [cell.value for cell in base_sheet[1]]
        try:
            value_idx = headers.index("值")
            node_name_idx = headers.index("节点名称")
            node_attr_idx = headers.index("节点属性")
            model_id_idx = headers.index("编号")
            trigger_idx = headers.index("触发词")
            img_path_idx = headers.index("图片路径")
        except ValueError as e:
            logger.error(f"底模工作簿缺少必要的列: {e}")
            return
        
        # 处理每一行底模
        for row_idx, row in enumerate(base_sheet.iter_rows(min_row=2), start=2):
            try:
                model_path = row[value_idx].value
                if not model_path:
                    continue
                
                node_name = row[node_name_idx].value
                node_attr = row[node_attr_idx].value
                model_id = row[model_id_idx].value
                trigger_word = row[trigger_idx].value
                img_path = row[img_path_idx].value
                
                # 验证必要参数
                if not all([node_name, node_attr]):
                    logger.warning(f"第{row_idx}行缺少节点名称或节点属性")
                    continue
                
                # 处理模型路径可能是数组的情况
                if isinstance(model_path, list):
                    # 如果是数组，取第一个元素作为模型路径用于查找
                    model_path_for_search = os.path.normpath(model_path[0])
                else:
                    model_path_for_search = os.path.normpath(model_path)
                
                # 如果编号和触发词为空，从模型信息表中查找
                if not model_id or not trigger_word:
                    for sd_model in self.sd_models:
                        if model_path_for_search and sd_model.get("文件名"):
                            # 规范化比较路径
                            sd_path = os.path.normpath(sd_model.get("文件名"))
                            if model_path_for_search in sd_path or sd_path in model_path_for_search:
                                if not model_id:
                                    model_id = sd_model.get("编号")
                                    base_sheet.cell(row=row_idx, column=model_id_idx+1, value=model_id)
                                if not trigger_word:
                                    trigger_word = sd_model.get("触发词")
                                    base_sheet.cell(row=row_idx, column=trigger_idx+1, value=trigger_word)
                                break
                
                # 构建图片保存路径
                prompt_id = params.get("提示词编号", {}).get("值")
                if not prompt_id:
                    logger.warning("缺少提示词编号参数")
                    continue
                    
                if model_id:
                    img_filename = f"{prompt_id}_{model_id}.png"
                    full_img_path = os.path.normpath(os.path.join(save_path, img_filename))
                    
                    # 检查图片是否已存在
                    if not img_path and not os.path.exists(full_img_path):
                        try:
                            # 创建workflow的副本，确保每次请求都使用新的workflow实例
                            import copy
                            workflow_copy = copy.deepcopy(workflow)
                            
                            # 设置模型参数，处理可能是数组的情况
                            workflow_copy.set_node_param(node_name, node_attr, model_path)
                            
                            # 注入测试表格中的其他参数
                            self.inject_test_params(workbook, workflow_copy)
                            
                            # 如果有触发词，添加到正面提示词中
                            if trigger_word:
                                self.add_trigger_to_prompts(workflow_copy, params, trigger_word)
                            
                            # 生成图片
                            img_data = self.generate_image(workflow_copy)
                            if img_data:
                                # 保存图片
                                with PILImage.open(io.BytesIO(img_data)) as pil_img:
                                    pil_img = pil_img.convert("RGBA")
                                    pil_img.save(full_img_path, "PNG")
                                    logger.info(f"图片已保存: {full_img_path}")
                                    # 更新图片路径
                                    base_sheet.cell(row=row_idx, column=img_path_idx+1, value=full_img_path)
                        except Exception as e:
                            logger.error(f"处理第{row_idx}行时发生错误: {e}")
                    elif img_path:
                        logger.info(f"图片已存在: {img_path}")
                    elif os.path.exists(full_img_path):
                        logger.info(f"图片已存在: {full_img_path}")
                        # 更新图片路径
                        base_sheet.cell(row=row_idx, column=img_path_idx+1, value=full_img_path)
            except Exception as e:
                logger.error(f"处理第{row_idx}行时发生错误: {e}")
                continue
    
    def process_lora_models(self, workbook, workflow, params, save_path, lora_sheets):
        """
        处理Lora工作簿
        """
        # 获取默认底模信息
        default_model = params.get("默认底模", {}).get("值")
        default_model_node = params.get("默认底模", {}).get("节点名称")
        default_model_attr = params.get("默认底模", {}).get("节点属性")
        
        # 获取默认底模编号（不再使用'default'）
        default_model_id = None
        if isinstance(default_model, list) and len(default_model) > 0:
            # 如果默认底模是数组，使用第一个元素作为查找依据
            model_path_for_search = os.path.normpath(default_model[0])
        else:
            model_path_for_search = os.path.normpath(default_model)
            
        # 从模型信息表中查找默认底模的编号
        for sd_model in self.sd_models:
            if sd_model.get("文件名"):
                sd_path = os.path.normpath(sd_model.get("文件名"))
                if model_path_for_search in sd_path or sd_path in model_path_for_search:
                    default_model_id = sd_model.get("编号")
                    break
        
        if not default_model_id:
            logger.warning("未找到默认底模的编号，将使用文件名作为编号")
            if isinstance(default_model, list):
                default_model_id = os.path.basename(default_model[0]).split(".")[0]
            else:
                default_model_id = os.path.basename(default_model).split(".")[0]
        
        for lora_sheet_name in lora_sheets:
            lora_sheet = workbook[lora_sheet_name]
            lora_num = lora_sheet_name.split("-")[1]  # 获取Lora编号
            
            # 获取列索引
            headers = [cell.value for cell in lora_sheet[1]]
            
            # 新的列结构：值、节点名称、lora强度、clip强度、编号、触发词、图片路径、展示图片
            value_idx = headers.index("值") if "值" in headers else None
            node_name_idx = headers.index("节点名称") if "节点名称" in headers else None
            lora_strength_idx = headers.index("lora强度") if "lora强度" in headers else None
            clip_strength_idx = headers.index("clip强度") if "clip强度" in headers else None
            model_id_idx = headers.index("编号") if "编号" in headers else None
            trigger_idx = headers.index("触发词") if "触发词" in headers else None
            img_path_idx = headers.index("图片路径") if "图片路径" in headers else None
            img_display_idx = headers.index("展示图片") if "展示图片" in headers else None
            
            # 检查必要的列是否存在
            if any(idx is None for idx in [value_idx, node_name_idx]):
                logger.error(f"Lora-{lora_num}工作簿缺少必要的列: 值、节点名称")
                continue
            
            # 处理每一行Lora
            for row_idx, row in enumerate(lora_sheet.iter_rows(min_row=2), start=2):
                # 获取基本参数
                lora_value = row[value_idx].value
                node_name = row[node_name_idx].value
                lora_strength = row[lora_strength_idx].value if lora_strength_idx is not None else 1.0
                clip_strength = row[clip_strength_idx].value if clip_strength_idx is not None else 1.0
                
                # 如果基本参数不存在，跳过
                if not all([lora_value, node_name]):
                    continue
                
                # 获取其他参数
                lora_id = row[model_id_idx].value if model_id_idx is not None else None
                trigger_word = row[trigger_idx].value if trigger_idx is not None else None
                img_path = row[img_path_idx].value if img_path_idx is not None else None
                
                # 解析lora_value可能是数组的情况
                lora_value = self.parse_value(lora_value)
                
                # 如果编号和触发词为空，从模型信息表中查找
                if not lora_id or not trigger_word:
                    lora_path_for_search = lora_value
                    if isinstance(lora_value, list) and len(lora_value) > 0:
                        lora_path_for_search = lora_value[0]
                    
                    for lora_model in self.lora_models:
                        if (lora_path_for_search and 
                            lora_model.get("文件名") and 
                            lora_model.get("文件名") in lora_path_for_search):
                            if not lora_id:
                                lora_id = lora_model.get("编号")
                                if model_id_idx is not None:
                                    lora_sheet.cell(row=row_idx, column=model_id_idx+1, value=lora_id)
                            if not trigger_word:
                                trigger_word = lora_model.get("触发词")
                                if trigger_idx is not None:
                                    lora_sheet.cell(row=row_idx, column=trigger_idx+1, value=trigger_word)
                            break
                
                # 构建图片保存路径
                prompt_id = params.get("提示词编号", {}).get("值")
                if lora_id:
                    # 构建图片文件名，使用实际的底模编号而非'default'
                    img_filename = f"{prompt_id}_{default_model_id}_{lora_id}&{lora_strength}.png"
                    full_img_path = os.path.normpath(os.path.join(save_path, img_filename))
                    
                    # 检查图片是否已存在
                    if not img_path and not os.path.exists(full_img_path):
                        try:
                            # 创建workflow的副本，确保每次请求都使用新的workflow实例
                            import copy
                            workflow_copy = copy.deepcopy(workflow)
                            
                            # 使用默认底模
                            workflow_copy.set_node_param(default_model_node, default_model_attr, default_model)
                            
                            # 处理Lora参数
                            # 1. 开启Lora开关
                            workflow_copy.set_node_param(node_name, "switch", "On")
                            
                            # 2. 设置Lora名称
                            workflow_copy.set_node_param(node_name, "lora_name", lora_value)
                            
                            # 3. 设置模型强度
                            workflow_copy.set_node_param(node_name, "strength_model", lora_strength)
                            
                            # 4. 设置clip强度
                            workflow_copy.set_node_param(node_name, "strength_clip", clip_strength)
                            
                            # 注入测试表格中的其他参数
                            self.inject_test_params(workbook, workflow_copy)
                            
                            # 如果有触发词，添加到正面提示词中
                            if trigger_word:
                                self.add_trigger_to_prompts(workflow_copy, params, trigger_word)
                            
                            # 生成图片
                            img_data = self.generate_image(workflow_copy)
                            if img_data:
                                # 保存图片
                                with PILImage.open(io.BytesIO(img_data)) as pil_img:
                                    pil_img = pil_img.convert("RGBA")
                                    pil_img.save(full_img_path, "PNG")
                                    logger.info(f"图片已保存: {full_img_path}")
                                    # 更新图片路径
                                    if img_path_idx is not None:
                                        lora_sheet.cell(row=row_idx, column=img_path_idx+1, value=full_img_path)
                        except Exception as e:
                            logger.error(f"处理Lora时发生错误: {e}")
                    elif img_path:
                        logger.info(f"图片已存在: {img_path}")
                    elif os.path.exists(full_img_path):
                        logger.info(f"图片已存在: {full_img_path}")
                        # 更新图片路径
                        if img_path_idx is not None:
                            lora_sheet.cell(row=row_idx, column=img_path_idx+1, value=full_img_path)
    
    def process_test_combinations(self, workbook, workflow, params, save_path):
        """
        处理测试工作簿中的组合测试
        """
        # 获取底模数据
        base_models = []
        base_sheet = workbook["底模"]
        base_headers = [cell.value for cell in base_sheet[1]]
        value_idx = base_headers.index("值") if "值" in base_headers else None
        node_name_idx = base_headers.index("节点名称") if "节点名称" in base_headers else None
        node_attr_idx = base_headers.index("节点属性") if "节点属性" in base_headers else None
        model_id_idx = base_headers.index("编号") if "编号" in base_headers else None
        trigger_idx = base_headers.index("触发词") if "触发词" in base_headers else None
        
        if all([value_idx is not None, node_name_idx is not None, node_attr_idx is not None, model_id_idx is not None]):
            for row in base_sheet.iter_rows(min_row=2, values_only=True):
                if row[value_idx]:  # 值不为空
                    base_models.append({
                        "值": row[value_idx],
                        "节点名称": row[node_name_idx],
                        "节点属性": row[node_attr_idx],
                        "编号": row[model_id_idx],
                        "触发词": row[trigger_idx] if trigger_idx is not None and row[trigger_idx] else ""
                    })
        
        # 获取Lora数据
        lora_data = {}
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("Lora-"):
                lora_sheet = workbook[sheet_name]
                lora_num = sheet_name.split("-")[1]
                lora_headers = [cell.value for cell in lora_sheet[1]]
                
                # 获取新的列结构索引：值、节点名称、lora强度、clip强度、编号、触发词、图片路径、展示图片
                value_idx = lora_headers.index("值") if "值" in lora_headers else None
                node_name_idx = lora_headers.index("节点名称") if "节点名称" in lora_headers else None
                lora_strength_idx = lora_headers.index("lora强度") if "lora强度" in lora_headers else None
                clip_strength_idx = lora_headers.index("clip强度") if "clip强度" in lora_headers else None
                model_id_idx = lora_headers.index("编号") if "编号" in lora_headers else None
                trigger_idx = lora_headers.index("触发词") if "触发词" in lora_headers else None
                
                # 检查必要的列是否存在
                if any(idx is None for idx in [value_idx, node_name_idx, model_id_idx]):
                    logger.error(f"Lora-{lora_num}工作簿缺少必要的列: 值、节点名称、编号")
                    continue
                
                lora_models = []
                for row in lora_sheet.iter_rows(min_row=2, values_only=True):
                    # 获取基本参数
                    lora_value = row[value_idx]
                    node_name = row[node_name_idx]
                    lora_strength = row[lora_strength_idx] if lora_strength_idx is not None and row[lora_strength_idx] else 1.0
                    clip_strength = row[clip_strength_idx] if clip_strength_idx is not None and row[clip_strength_idx] else 1.0
                    model_id = row[model_id_idx]
                    trigger_word = row[trigger_idx] if trigger_idx is not None and row[trigger_idx] else ""
                    
                    # 如果基本参数不存在，跳过
                    if not all([lora_value, node_name, model_id]):
                        continue
                    
                    # 创建Lora模型对象
                    lora_model = {
                        "编号": model_id,
                        "值": lora_value,
                        "节点名称": node_name,
                        "lora强度": lora_strength,
                        "clip强度": clip_strength,
                        "触发词": trigger_word
                    }
                    lora_models.append(lora_model)
                
                if lora_models:
                    lora_data[lora_num] = lora_models
        
        # 如果没有测试工作簿，创建一个
        if "测试" not in workbook.sheetnames:
            test_sheet = workbook.create_sheet("测试")
            # 创建表头
            headers = ["底模编号", "底模名称"]
            for lora_num in sorted(lora_data.keys()):
                headers.extend([f"Lora-{lora_num}编号", f"Lora-{lora_num}名称", f"Lora-{lora_num}强度"])
            headers.extend(["图片路径", "展示图片"])
            for col_idx, header in enumerate(headers, start=1):
                test_sheet.cell(row=1, column=col_idx, value=header)
        else:
            test_sheet = workbook["测试"]
        
        # 生成所有可能的组合
        combinations = self.generate_combinations(base_models, lora_data)
        
        # 处理每个组合
        for row_idx, combo in enumerate(combinations, start=2):
            # 检查该行是否已存在
            if row_idx <= test_sheet.max_row:
                existing_img_path = test_sheet.cell(row=row_idx, column=len(headers)-1).value
                if existing_img_path:
                    logger.info(f"组合已测试: {existing_img_path}")
                    continue
            
            # 写入组合信息
            for col_idx, value in enumerate(combo[:-1], start=1):
                test_sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # 构建图片文件名
            prompt_id = params.get("提示词编号", {}).get("值")
            base_id = combo[0]  # 底模编号
            
            # 收集Lora信息
            lora_info = []
            col_offset = 2
            for lora_num in sorted(lora_data.keys()):
                if col_offset + 2 < len(combo):
                    lora_id = combo[col_offset]      # Lora编号
                    lora_strength = combo[col_offset+2]  # Lora强度
                    if lora_id:
                        lora_info.append((lora_id, lora_strength))
                col_offset += 3
            
            # 构建图片文件名
            lora_part = "_".join([f"{lid}&{ls}" for lid, ls in lora_info]) if lora_info else ""
            img_filename = f"{prompt_id}_{base_id}{('_' + lora_part) if lora_part else ''}.png"
            full_img_path = os.path.normpath(os.path.join(save_path, img_filename))
            
            # 检查图片是否已存在
            if not os.path.exists(full_img_path):
                # 检查不同顺序的Lora组合
                if len(lora_info) >= 2:
                    # 尝试所有可能的Lora顺序组合
                    import itertools
                    found = False
                    for perm in itertools.permutations(lora_info):
                        lora_part_alt = "_".join([f"{lid}&{ls}" for lid, ls in perm])
                        alt_filename = f"{prompt_id}_{base_id}_{lora_part_alt}.png"
                        alt_path = os.path.normpath(os.path.join(save_path, alt_filename))
                        if os.path.exists(alt_path):
                            logger.info(f"找到不同顺序的图片: {alt_path}")
                            test_sheet.cell(row=row_idx, column=len(headers)-1, value=alt_path)
                            found = True
                            break
                    if found:
                        continue
                
                try:
                    # 创建workflow的副本，确保每次请求都使用新的workflow实例
                    import copy
                    workflow_copy = copy.deepcopy(workflow)
                    
                    # 注入测试表格中的其他参数
                    self.inject_test_params(workbook, workflow_copy)
                    
                    # 设置底模
                    for bm in base_models:
                        if bm["编号"] == base_id:
                            workflow_copy.set_node_param(bm["节点名称"], bm["节点属性"], bm["值"])
                            # 添加底模触发词
                            if bm["触发词"]:
                                self.add_trigger_to_prompts(workflow_copy, params, bm["触发词"])
                            break
                        
                    # 设置Lora
                    col_offset = 2
                    for lora_num in sorted(lora_data.keys()):
                        if col_offset + 2 < len(combo):
                            lora_id = combo[col_offset]      # Lora编号
                            lora_value = combo[col_offset+1]  # Lora值
                            lora_strength = combo[col_offset+2]  # Lora强度
                            
                            if lora_id and lora_value:
                                for lora in lora_data[lora_num]:
                                    if lora["编号"] == lora_id:
                                        # 使用新的列结构设置Lora参数
                                        node_name = lora["节点名称"]
                                        
                                        # 1. 开启Lora开关
                                        workflow_copy.set_node_param(node_name, "switch", "On")
                                        
                                        # 2. 设置Lora名称
                                        workflow_copy.set_node_param(node_name, "lora_name", lora_value)
                                        
                                        # 3. 设置模型强度
                                        workflow_copy.set_node_param(node_name, "strength_model", lora_strength)
                                        
                                        # 4. 设置clip强度
                                        clip_strength = lora.get("clip强度", lora_strength)  # 如果没有clip强度，使用lora强度
                                        workflow_copy.set_node_param(node_name, "strength_clip", clip_strength)
                                        
                                        # 添加触发词
                                        if lora["触发词"]:
                                            self.add_trigger_to_prompts(workflow_copy, params, lora["触发词"])
                                        break
                        col_offset += 3
                    
                    # 生成图片
                    img_data = self.generate_image(workflow_copy)
                    if img_data:
                        # 保存图片
                        with PILImage.open(io.BytesIO(img_data)) as pil_img:
                            pil_img = pil_img.convert("RGBA")
                            pil_img.save(full_img_path, "PNG")
                            logger.info(f"图片已保存: {full_img_path}")
                            # 更新图片路径
                            test_sheet.cell(row=row_idx, column=len(headers)-1, value=full_img_path)
                except Exception as e:
                    logger.error(f"处理组合测试时发生错误: {e}")
            else:
                logger.info(f"图片已存在: {full_img_path}")
                # 更新图片路径
                test_sheet.cell(row=row_idx, column=len(headers)-1, value=full_img_path)
    
    def generate_combinations(self, base_models, lora_data):
        """
        生成所有可能的组合
        """
        combinations = []
        
        # 对于每个底模
        for base_model in base_models:
            if not base_model["编号"]:
                continue
                
            # 基本组合：只有底模
            base_combo = [base_model["编号"], base_model["值"]]
            for lora_num in sorted(lora_data.keys()):
                base_combo.extend(["", "", ""])  # 为每个Lora位置添加空值（编号、值、强度）
            base_combo.append("")  # 图片路径
            combinations.append(base_combo)
            
            # 添加单个Lora的组合
            for lora_num, loras in lora_data.items():
                for lora in loras:
                    if not lora["编号"]:
                        continue
                    
                    combo = [base_model["编号"], base_model["值"]]
                    for curr_lora_num in sorted(lora_data.keys()):
                        if curr_lora_num == lora_num:
                            # 使用新的列结构：值、lora强度、clip强度
                            # 只考虑值、lora强度、clip强度这三个关键参数
                            lora_strength = lora.get("lora强度", 1.0)
                            combo.extend([lora["编号"], lora["值"], lora_strength])
                        else:
                            combo.extend(["", "", ""])
                    combo.append("")  # 图片路径
                    combinations.append(combo)
            
            # 添加多个Lora的组合（最多支持4个Lora）
            if len(lora_data) >= 2:
                # 生成2个Lora的组合
                self.add_multi_lora_combinations(combinations, base_model, lora_data, 2)
            
            if len(lora_data) >= 3:
                # 生成3个Lora的组合
                self.add_multi_lora_combinations(combinations, base_model, lora_data, 3)
                
            if len(lora_data) >= 4:
                # 生成4个Lora的组合
                self.add_multi_lora_combinations(combinations, base_model, lora_data, 4)
        
        return combinations
    
    def add_multi_lora_combinations(self, combinations, base_model, lora_data, num_loras):
        """
        添加多个Lora的组合
        """
        import itertools
        
        # 获取所有Lora编号
        lora_nums = sorted(lora_data.keys())
        
        # 生成所有可能的Lora编号组合
        for lora_num_combo in itertools.combinations(lora_nums, num_loras):
            # 对于每个Lora编号组合，生成所有可能的Lora模型组合
            lora_options = [lora_data[num] for num in lora_num_combo]
            for lora_combo in itertools.product(*lora_options):
                combo = [base_model["编号"], base_model["值"]]
                
                # 填充Lora信息
                lora_info = {}
                for idx, lora in enumerate(lora_combo):
                    lora_num = lora_num_combo[idx]
                    lora_info[lora_num] = lora
                
                for curr_lora_num in sorted(lora_data.keys()):
                    if curr_lora_num in lora_info:
                        lora = lora_info[curr_lora_num]
                        # 使用新的列结构：值、lora强度、clip强度
                        # 只考虑值、lora强度、clip强度这三个关键参数
                        lora_strength = lora.get("lora强度", 1.0)
                        combo.extend([lora["编号"], lora["值"], lora_strength])
                    else:
                        combo.extend(["", "", ""])
                
                combo.append("")  # 图片路径
                combinations.append(combo)
    
    def add_trigger_to_prompts(self, workflow, params, trigger_word):
        """
        将触发词添加到正面提示词中
        """
        # 查找所有正面提示词参数
        for param_name, param_data in params.items():
            if "正面" in param_name and param_data.get("节点名称") and param_data.get("节点属性"):
                current_prompt = workflow.get_node_param(param_data["节点名称"], param_data["节点属性"])
                # 在提示词前面添加触发词
                if trigger_word and current_prompt:
                    new_prompt = f"{trigger_word}. {current_prompt}"
                    # 将修改后的提示词写入到workflow中对应的节点参数
                    logger.info(f"添加触发词到提示词: {param_data['节点名称']} > {param_data['节点属性']} > {trigger_word}")
                    workflow.set_node_param(param_data["节点名称"], param_data["节点属性"], new_prompt)
    
    def generate_image(self, workflow):
        """
        使用ComfyUI生成图片
        
        Args:
            workflow: 工作流对象
            
        Returns:
            bytes: 图片数据，如果生成失败则返回None
        """
        try:
            # 延迟初始化WebSocket客户端，只在需要生成图片时才连接
            if self.ws_client is None:
                logger.info(f"正在连接到ComfyUI服务器: {self.comfy_host}")
                self.ws_client = ComfyWebSocketClient(self.comfy_host)
            
            # 获取图片
            images = self.ws_client.get_images(workflow)
            
            # 遍历所有节点及其对应的图片数据
            for node_id, img_list in images.items():
                for img_idx, img_data in enumerate(img_list):
                    # 只返回第一张图片
                    return img_data
            
            return None
        except Exception as e:
            logger.error(f"生成图片失败: {e}")
            return None
            
    def inject_test_params(self, workbook, workflow):
        """
        注入测试表格中的参数
        
        Args:
            workbook: 测试表格工作簿
            workflow: 工作流对象
        """
        sheet = workbook["参数"]
        headers = [cell.value for cell in sheet[1]]
        
        # 查找值、节点名称、节点属性列
        name_indices = [i for i, h in enumerate(headers) if h == "名称"]
        value_indices = [i for i, h in enumerate(headers) if h == "值" ]
        node_name_indices = [i for i, h in enumerate(headers) if h == "节点名称"]
        node_attr_indices = [i for i, h in enumerate(headers) if h == "节点属性"]
        
        # 遍历所有行
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            # 检查每组值、节点名称、节点属性
            for val_idx, name_idx, attr_idx in zip(value_indices, node_name_indices, node_attr_indices):
                name = row[name_idx].value
                value = row[val_idx].value
                node_name = row[name_idx].value
                node_attr = row[attr_idx].value
                
                # 确保三个值都不为空，才注入参数
                if value is not None and node_name and node_attr:
                    if '正面' in node_name:
                        continue
                    if '默认底模' in node_name:
                        continue
                    try:
                        # 解析值可能是数组的情况
                        parsed_value = self.parse_value(value)
                        logger.info(f"注入参数: {node_name} > {node_attr} > {parsed_value}")
                        workflow.set_node_param(node_name, node_attr, parsed_value)
                    except Exception as e:
                        logger.error(f"注入参数失败: {e}")
    
    def reinsert_image(self, file_path):
        """
        将图片插入到Excel表格中
        
        Args:
            file_path: Excel文件路径
        """
        try:
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.styles import Alignment
            from openpyxl.utils import get_column_letter
            
            # 加载工作簿
            wb = load_workbook(file_path, data_only=True)
            
            # 先遍历所有工作表清空已有图片
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 获取图片的列表
                images = list(ws._images)  # 创建一个副本以避免在迭代时修改列表长度
                for drawing in images:
                    ws._images.remove(drawing)
            
            # 遍历所有工作表插入图片
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2):
                    if row[0].value is None:
                        continue
                    for col_index, cell in enumerate(row, start=1):
                        value = cell.value
                        if value is None or not isinstance(value, str):
                            continue
                        if os.path.exists(value) and value.lower().endswith('.png'):
                            try:
                                # 直接打开原始图片，不保存压缩后的图片
                                img = PILImage.open(value)
                                # 获取原始图片的宽高
                                width, height = img.size
                                # 只限制最大高度，不限制宽度
                                max_height = 512
                                # 如果高度超过最大高度，按高度等比例缩放
                                if height > max_height:
                                    ratio = max_height / height
                                    new_width = int(width * ratio)
                                    new_height = max_height
                                    # 进行缩放
                                    img = img.resize((new_width, new_height), PILImage.LANCZOS)
                            except Exception as e:
                                logger.error(f"处理图片 {value} 时出错: {e}")
                                continue
                            
                            try:
                                # 使用BytesIO创建内存中的图片对象，不需要保存到磁盘
                                img_buffer = io.BytesIO()
                                img.save(img_buffer, format='PNG')
                                img_buffer.seek(0)
                                img = OpenpyxlImage(img_buffer)
                                
                                # 调整单元格宽高
                                img_width = img.width
                                img_height = img.height
                                img_scale = img_width / img_height
                                
                                # 固定行高为100磅
                                height_pt = 100  # 目标高度为 100 pt
                                height_px = height_pt * (4 / 3)  # 转换为像素
                                
                                # 只限制高度，不限制宽度
                                if img_height > height_px:
                                    # 按高度缩放
                                    new_height = height_px
                                    new_width = int(new_height * img_scale)  # 保持宽高比
                                else:
                                    # 图片高度小于行高，保持原始大小
                                    new_width = img_width
                                    new_height = img_height
                                
                                # 设置图片的新宽度和高度
                                img.width = new_width
                                img.height = new_height
                                
                                # 根据图片宽度计算列宽（Excel列宽单位为字符，约等于像素/8）
                                width_ch = new_width / 8
                                
                                # 设置单元格宽高并插入图片
                                colname = get_column_letter(col_index+1)
                                rowindex = cell.row
                                ws.column_dimensions[colname].width = width_ch  # 根据图片宽度动态设置列宽
                                ws.row_dimensions[rowindex].height = height_pt  # 行高固定为100磅
                                ws.add_image(img, f"{colname}{rowindex}")
                                logger.info(f'插入图片: {value} -> 单元格 {colname}{rowindex}，列宽: {width_ch:.2f}字符')
                            except Exception as e:
                                logger.error(f"插入图片 {value} 时出错: {e}")
            
            # 设置单元格自动换行
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=1):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
            
            # 保存工作簿
            wb.save(file_path)
            logger.info(f"图片插入完成，已保存到 {file_path}")
        except Exception as e:
            logger.error(f"插入图片时发生错误: {e}")


def main():
    """
    主函数入口，直接使用默认参数运行批量测试
    """
    # 设置默认参数
    model_info_path = "E:\\models\\model_info.xlsx"  # 模型信息表格路径
    test_file_path = "D:\\Code\\MY_ComfyUI\\# 模型测试\\#5 提示词单次测试-Unet.xlsx"  # 测试文件路径
    comfy_host = "127.0.0.1:8191"  # ComfyUI服务器地址
    insert_image = True  # 是否插入图片到Excel表格中
    
    # 创建批量测试模型对象
    logger.info(f"初始化批量测试模型，模型信息: {model_info_path}, ComfyUI服务器: {comfy_host}")
    tester = BatchModelTester(model_info_path, comfy_host)
    
    # 加载模型信息
    logger.info("加载模型信息...")
    tester.load_model_info()
    
    # 处理测试文件
    logger.info(f"开始处理测试文件: {test_file_path}")
    tester.process_test_file(test_file_path)
    
    # 插入图片到Excel表格中
    if insert_image:
        logger.info("开始将图片插入到Excel表格中...")
        tester.reinsert_image(test_file_path)
        
    logger.info("批量测试完成！")


if __name__ == "__main__":
    main()