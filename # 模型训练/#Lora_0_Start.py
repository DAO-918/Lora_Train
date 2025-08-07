import os
import sys
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
import datetime
import shutil
import subprocess
import importlib.util
import importlib.machinery
import inspect
import yaml
import pathlib

# 定义训练集类型和地址
CHARACTER_DIR = "E:\\Design\\Character"
STYLES_DIR = "E:\\Design\\Styles"

def parse_arguments():
    """
    解析命令行参数
    """
    parser = argparse.ArgumentParser(description="Lora训练初始化脚本")
    parser.add_argument("--project_path", type=str, help="项目路径，如果不提供，将自动检测")
    parser.add_argument("--bat_dir", action="store_true", help="使用bat文件所在目录作为项目路径")
    parser.add_argument("--use_yaml", action="store_true", help="使用yaml配置文件中的项目路径")
    return parser.parse_args()

def detect_project_type_and_name(project_path=None):
    """
    检测项目类型和名称
    如果提供了项目路径，直接从中提取信息
    否则检查当前工作目录是否在训练集目录下
    """
    print("\n===== 开始检测项目类型和名称 =====")
    print(f"输入的项目路径: {project_path if project_path else '无，将使用当前工作目录'}")
    
    if project_path:
        if os.path.exists(project_path):
            # 从提供的路径中提取项目类型和名称
            if project_path.startswith(CHARACTER_DIR):
                project_type = "Character"
                project_name = os.path.basename(project_path)
            elif project_path.startswith(STYLES_DIR):
                project_type = "Styles"
                project_name = os.path.basename(project_path)
            else:
                # 检查是否是有效的项目路径（包含必要的文件或目录结构）
                # 这里我们简单地检查路径是否存在，并假设它是一个有效的项目路径
                # 用户可以通过bat文件传递任意路径作为项目路径
                project_type = "Custom"
                project_name = os.path.basename(project_path)
                print(f"警告: 提供的路径 {project_path} 不在标准训练集目录下，但将作为自定义项目路径使用")
                return project_type, project_name, project_path
            # 如果是标准目录，也直接返回结果
            print(f"项目类型识别成功: {project_type}")
            print(f"项目名称: {project_name}")
            print(f"项目完整路径: {project_path}")
            print("===== 项目类型和名称检测完成 =====\n")
            return project_type, project_name, project_path
    
    # 如果没有提供路径，检查当前工作目录
    current_dir = os.getcwd()
    
    # 检查是否在角色训练集目录下
    if current_dir.startswith(CHARACTER_DIR):
        project_type = "Character"
        project_name = os.path.basename(current_dir)
        print(f"项目类型识别成功: {project_type}")
        print(f"项目名称: {project_name}")
        print(f"项目完整路径: {current_dir}")
        print("===== 项目类型和名称检测完成 =====\n")
        return project_type, project_name, current_dir
    
    # 检查是否在风格训练集目录下
    elif current_dir.startswith(STYLES_DIR):
        project_type = "Styles"
        project_name = os.path.basename(current_dir)
        print(f"项目类型识别成功: {project_type}")
        print(f"项目名称: {project_name}")
        print(f"项目完整路径: {current_dir}")
        print("===== 项目类型和名称检测完成 =====\n")
        return project_type, project_name, current_dir
    
    # 如果都不是，提示错误
    raise ValueError(f"当前目录 {current_dir} 不在有效的训练集目录下，请切换到训练集目录或提供有效的项目路径")

def create_training_info_excel(project_type, project_name, project_path):
    """
    创建训练信息Excel文件
    """
    print("\n===== 开始创建训练信息Excel文件 =====")
    print(f"项目类型: {project_type}")
    print(f"项目名称: {project_name}")
    print(f"项目路径: {project_path}")
    
    excel_path = os.path.join(project_path, "训练信息.xlsx")
    
    # 直接覆盖已存在的文件，不创建备份
    
    # 创建新的工作簿
    wb = Workbook()
    
    # 创建步骤工作表
    steps_sheet = wb.active
    # 设置工作表名称为"步骤"
    if steps_sheet is not None:
        steps_sheet.title = "步骤"
    else:
        print("警告: 工作表对象为None，无法设置标题")
    
    # 设置列宽
    steps_sheet.column_dimensions['A'].width = 20
    steps_sheet.column_dimensions['B'].width = 15
    steps_sheet.column_dimensions['C'].width = 20
    steps_sheet.column_dimensions['D'].width = 30
    
    # 添加标题行
    steps_sheet['A1'] = "步骤名称"
    steps_sheet['B1'] = "是否执行"
    steps_sheet['C1'] = "完成时间"
    steps_sheet['D1'] = "完成结果"
    
    # 设置标题行样式
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)
    for cell in steps_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加步骤数据
    steps = [
        "图片尺寸标准化",
        "图片描述生成",
        "图片描述优化",
        "图片描述插入",
        "模型Lora训练",
        "模型Lora测试",
        "版本号",
        "反推提示词",
        "触发词",
        "插入内容",
        "提示词编号",
        "训练模板",
        "是否关机"
    ]
    
    # 设置默认值
    default_values = {
        "图片尺寸标准化": 0,
        "图片描述生成": 0,
        "图片描述优化": 0,
        "图片描述插入": 0,
        "模型Lora训练": 0,
        "模型Lora测试": 0,
        "版本号": "v1",
        "反推提示词": "",  # 将根据项目类型设置
        "触发词": project_name,  # 设置为项目名称
        "插入内容": f"{project_name}. ", 
        "提示词编号": "com01",
        "训练模板": "16epoch-1-1024-batch=1_3e-4",
        "是否关机": 0
    }
    
    # 根据项目类型设置反推提示词
    if project_type == "Styles":
        default_values["反推提示词"] = ""
    elif project_type == "Character":
        default_values["反推提示词"] = ""
    else:  # 其他项目类型
        default_values["反推提示词"] = "通用提示词"
    
    for i, step in enumerate(steps, 2):
        steps_sheet[f'A{i}'] = step
        
        # 设置默认值
        if step in default_values:
            if isinstance(default_values[step], int):
                steps_sheet[f'B{i}'] = default_values[step]
            else:
                steps_sheet[f'B{i}'] = default_values[step]
    
    # 创建提示词工作表
    prompts_sheet = wb.create_sheet(title="提示词")
    
    # 设置列宽
    prompts_sheet.column_dimensions['A'].width = 50
    prompts_sheet.column_dimensions['B'].width = 50
    prompts_sheet.column_dimensions['C'].width = 20
    prompts_sheet.column_dimensions['D'].width = 10
    
    # 添加标题行
    prompts_sheet['A1'] = "英文提示词"
    prompts_sheet['B1'] = "中文提示词"
    prompts_sheet['C1'] = "图片预览"
    prompts_sheet['D1'] = "图片路径"
    
    # 设置标题行样式
    for cell in prompts_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 设置行高
    for i in range(2, 21):  # 设置前20行的行高
        prompts_sheet.row_dimensions[i].height = 100
    
    # 保存工作簿
    wb.save(excel_path)
    print(f"已创建训练信息Excel文件: {excel_path}")
    print(f"创建了工作表: {', '.join([sheet.title for sheet in wb.worksheets])}")
    print(f"设置了默认值: 触发词={project_name}, 提示词编号=com01")
    print("===== 训练信息Excel文件创建完成 =====\n")
    
    return excel_path

def create_directory_structure(project_path):
    """
    创建必要的目录结构
    """
    print("\n===== 开始创建项目目录结构 =====")
    print(f"项目路径: {project_path}")
    
    # 创建resize目录
    resize_dir = os.path.join(project_path, "resize")
    os.makedirs(resize_dir, exist_ok=True)
    
    # 创建gemini目录
    gemini_dir = os.path.join(project_path, "gemini")
    os.makedirs(gemini_dir, exist_ok=True)
    
    print(f"已创建必要的目录结构:")
    print(f"  - resize目录: {resize_dir}")
    print(f"  - gemini目录: {gemini_dir}")
    print("===== 项目目录结构创建完成 =====\n")

def read_execution_flags(excel_path):
    """
    读取训练信息Excel文件中的执行标志
    
    Args:
        excel_path: Excel文件路径
        
    Returns:
        执行标志字典，键为步骤名称，值为是否执行的标志
    """
    try:
        wb = load_workbook(excel_path)
        steps_sheet = wb["步骤"]
        
        print("\n===== 开始读取执行标志 =====")
        print(f"Excel文件路径: {excel_path}")
        
        execution_flags = {}
        for row in range(2, steps_sheet.max_row + 1):
            step_name = steps_sheet.cell(row=row, column=1).value
            execution_flag = steps_sheet.cell(row=row, column=2).value
            execution_flags[step_name] = execution_flag
            print(f"步骤 '{step_name}' 的执行标志: {execution_flag}")
        
        print("===== 执行标志读取完成 =====\n")
        return execution_flags
    except Exception as e:
        print(f"读取执行标志时出错: {e}")
        return {}

def check_step_completed(excel_path, step_name):
    """
    检查步骤是否已成功完成
    
    Args:
        excel_path: Excel文件路径
        step_name: 步骤名称
        
    Returns:
        如果步骤已成功完成返回True，否则返回False
    """
    try:
        wb = load_workbook(excel_path)
        steps_sheet = wb["步骤"]
        
        # 查找步骤所在行
        step_row = None
        for row in range(2, steps_sheet.max_row + 1):
            if steps_sheet.cell(row=row, column=1).value == step_name:
                step_row = row
                break
        
        if step_row:
            # 检查完成结果列（第4列）
            result = steps_sheet.cell(row=step_row, column=4).value
            completed = result == "成功"
            print(f"检查步骤 '{step_name}' 完成状态: {'已完成' if completed else '未完成或失败'}")
            return completed
        else:
            print(f"警告: 在Excel中未找到步骤 {step_name}")
            return False
    except Exception as e:
        print(f"检查步骤完成状态时出错: {e}")
        return False

def update_step_result(excel_path, step_name, success):
    """
    更新训练信息Excel文件中的步骤执行结果
    
    Args:
        excel_path: Excel文件路径
        step_name: 步骤名称
        success: 是否成功
    """
    try:
        wb = load_workbook(excel_path)
        steps_sheet = wb["步骤"]
        
        # 查找步骤所在行
        step_row = None
        for row in range(2, steps_sheet.max_row + 1):
            if steps_sheet.cell(row=row, column=1).value == step_name:
                step_row = row
                break
        
        if step_row:
            # 更新完成时间
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            steps_sheet.cell(row=step_row, column=3).value = current_time
            
            # 更新完成结果
            result = "成功" if success else "失败"
            steps_sheet.cell(row=step_row, column=4).value = result
            
            # 保存工作簿
            wb.save(excel_path)
            print(f"已更新步骤 '{step_name}' 的执行结果: {result}")
            print(f"完成时间: {current_time}")
        else:
            print(f"警告: 在Excel中未找到步骤 {step_name}")
    except Exception as e:
        print(f"更新步骤执行结果时出错: {e}")

def get_script_path(script_name):
    """
    获取脚本的绝对路径
    
    Args:
        script_name: 脚本文件名
        
    Returns:
        脚本的绝对路径
    """
    # 获取当前脚本所在目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    script_path = os.path.join(current_dir, script_name)
    
    # 检查脚本是否存在
    if not os.path.exists(script_path):
        print(f"警告: 脚本 {script_name} 在当前目录 {current_dir} 中不存在")
        print(f"尝试在固定路径中查找脚本...")
        # 尝试使用绝对路径
        script_path = os.path.join(r"D:\Code\MY_ComfyUI\# 模型训练", script_name)
        if not os.path.exists(script_path):
            raise FileNotFoundError(f"无法找到脚本: {script_name}")
    
    return script_path

def run_script(script_name, args=None, script_description=None):
    """
    通用脚本运行函数
    
    Args:
        script_name: 脚本文件名
        args: 传递给脚本的参数列表
        script_description: 脚本描述，用于日志输出
    
    Returns:
        成功返回True，失败返回False
    """
    try:
        # 获取脚本路径
        script_path = get_script_path(script_name)
        
        # 如果没有提供描述，使用脚本名称
        if not script_description:
            script_description = script_name
        
        print(f"\n===== 开始运行{script_description}脚本 =====")
        print(f"脚本路径: {script_path}")
        if args:
            print(f"传递参数: {args}")
        
        # 优先使用importlib方式直接导入并执行模块
        try:
            print("使用importlib方法导入脚本...")
            # 使用importlib动态导入模块
            module_name = script_name.replace("#", "").replace(".py", "")
            loader = importlib.machinery.SourceFileLoader(module_name, script_path)
            spec = importlib.util.spec_from_loader(module_name, loader)
            script_module = importlib.util.module_from_spec(spec)
            loader.exec_module(script_module)
            
            # 保存当前sys.argv和标准输出
            old_argv = sys.argv.copy()
            
            # 修改sys.argv以传递参数
            sys.argv = [script_path]
            if args:
                if isinstance(args, list):
                    sys.argv.extend(args)
                else:
                    sys.argv.append(args)
                    
            # 调用模块的main函数
            if hasattr(script_module, 'main'):
                script_module.main()
                success = True
            else:
                print(f"警告: 脚本 {script_name} 没有main函数，尝试使用subprocess方法")
                raise AttributeError(f"脚本 {script_name} 没有main函数")
                
            # 恢复原始sys.argv
            sys.argv = old_argv
            
        except Exception as e:
            print(f"使用importlib方法执行脚本时出错: {e}")
            print("尝试使用subprocess方法执行脚本...")
            
            # 备选方法: 使用subprocess调用脚本
            cmd = [sys.executable, script_path]
            if args:
                if isinstance(args, list):
                    cmd.extend(args)
                else:
                    cmd.append(args)
                    
            result = subprocess.run(cmd, capture_output=True, text=True)
            print("脚本输出:")
            print(result.stdout)
            if result.stderr:
                print("错误输出:")
                print(result.stderr)
                
            success = result.returncode == 0
        
        print(f"{script_description}脚本执行{'成功' if success else '失败'}")
        print(f"===== {script_description}脚本执行完成 =====\n")
        return success
    except Exception as e:
        print(f"运行{script_description}脚本时出错: {e}")
        return False

def run_resize_script(resize_dir):
    """
    运行图片尺寸标准化脚本
    
    Args:
        resize_dir: resize文件夹路径
    
    Returns:
        成功返回True，失败返回False
    """
    # 确保resize_dir是绝对路径
    resize_dir = os.path.abspath(resize_dir)
    
    # 检查resize_dir是否存在
    if not os.path.exists(resize_dir):
        print(f"警告: resize目录 {resize_dir} 不存在，将创建该目录")
        print(f"创建目录: {resize_dir}")
        os.makedirs(resize_dir, exist_ok=True)
    
    return run_script("#Lora_1_图片尺寸-ARB桶.py", resize_dir, "图片尺寸标准化")

def run_image_description_script(resize_dir):
    """
    运行图片描述生成脚本
    
    Args:
        resize_dir: resize文件夹路径
    
    Returns:
        成功返回True，失败返回False
    """
    # 确保resize_dir是绝对路径
    resize_dir = os.path.abspath(resize_dir)
    
    return run_script("#Lora_2_画面描述-Gemini.py", resize_dir, "图片描述生成")

def run_description_optimization_script(project_path, ai_translation=False):
    """
    运行图片描述优化脚本
    
    Args:
        project_path: 项目路径
        ai_translation: 是否执行AI翻译
    
    Returns:
        成功返回True，失败返回False
    """
    # 确保project_path是绝对路径
    project_path = os.path.abspath(project_path)
    
    # 如果需要执行AI翻译，传递额外参数
    args = [project_path]
    if ai_translation:
        args.append("--translate")
    
    return run_script("#Lora_3_画面描述优化-Gemini.py", args, "图片描述优化")

def run_model_training_script(project_path):
    """
    运行模型Lora训练脚本
    
    Args:
        project_path: 项目路径
    
    Returns:
        成功返回True，失败返回False
    """
    # 确保project_path是绝对路径
    project_path = os.path.abspath(project_path)
    
    return run_script("#Lora_4_模型训练.py", project_path, "模型Lora训练")

def run_model_test_script(project_path):
    """
    运行模型Lora测试脚本
    
    Args:
        project_path: 项目路径
    
    Returns:
        成功返回True，失败返回False
    """
    # 确保project_path是绝对路径
    project_path = os.path.abspath(project_path)
    
    # 使用命名参数格式传递项目路径
    return run_script("#Lora_5_模型测试.py", ["--project_path", project_path], "模型Lora测试")

def run_description_insertion_script(project_path, excel_path):
    """
    图片描述插入功能
    在gemini目录下的所有txt文件头部插入Excel中定义的"插入内容"
    
    Args:
        project_path: 项目路径
        excel_path: Excel文件路径
    
    Returns:
        成功返回True，失败返回False
    """
    try:
        print("\n===== 开始执行图片描述插入 =====")
        print(f"项目路径: {project_path}")
        print(f"Excel文件路径: {excel_path}")
        
        # 确保project_path是绝对路径
        project_path = os.path.abspath(project_path)
        
        # 获取gemini目录路径
        gemini_dir = os.path.join(project_path, "gemini")
        print(f"gemini目录: {gemini_dir}")
        
        # 检查gemini目录是否存在
        if not os.path.exists(gemini_dir):
            print(f"警告: gemini目录 {gemini_dir} 不存在，将创建该目录")
            os.makedirs(gemini_dir, exist_ok=True)
        
        # 从Excel中读取插入内容
        wb = load_workbook(excel_path)
        steps_sheet = wb["步骤"]
        
        # 查找插入内容所在行
        insert_content = ""
        for row in range(2, steps_sheet.max_row + 1):
            if steps_sheet.cell(row=row, column=1).value == "插入内容":
                insert_content = steps_sheet.cell(row=row, column=2).value
                break
        
        if not insert_content:
            print("警告: 未在Excel中找到插入内容或插入内容为空")
            print("===== 图片描述插入执行失败 =====\n")
            return False
        
        print(f"从Excel中读取到的插入内容: {insert_content}")
        
        # 获取gemini目录下的所有txt文件
        txt_files = []
        for root, _, files in os.walk(gemini_dir):
            for file in files:
                if file.lower().endswith(".txt"):
                    txt_files.append(os.path.join(root, file))
        
        if not txt_files:
            print(f"警告: 在gemini目录 {gemini_dir} 中未找到任何txt文件")
            print("===== 图片描述插入执行失败 =====\n")
            return False
        
        print(f"找到 {len(txt_files)} 个txt文件需要处理")
        
        # 处理每个txt文件，在头部插入内容
        success_count = 0
        for txt_file in txt_files:
            try:
                # 读取文件内容
                with open(txt_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                # 检查文件是否已经包含插入内容
                if content.startswith(insert_content):
                    print(f"文件 {os.path.basename(txt_file)} 已包含插入内容，跳过")
                    success_count += 1
                    continue
                
                # 额外检查：去除空白字符后再次比较，以处理可能的空格或换行符差异
                if content.strip() and insert_content.strip() and content.strip().startswith(insert_content.strip()):
                    print(f"文件 {os.path.basename(txt_file)} 已包含插入内容（忽略空白字符），跳过")
                    success_count += 1
                    continue
                
                # 在头部插入内容
                new_content = insert_content + content
                
                # 写回文件
                with open(txt_file, 'w', encoding='utf-8') as f:
                    f.write(new_content)
                
                print(f"成功处理文件: {os.path.basename(txt_file)}")
                success_count += 1
            except Exception as e:
                print(f"处理文件 {os.path.basename(txt_file)} 时出错: {e}")
        
        success = success_count == len(txt_files)
        print(f"处理完成: {success_count}/{len(txt_files)} 个文件成功")
        print(f"===== 图片描述插入执行{'成功' if success else '失败'} =====\n")
        return success
    except Exception as e:
        print(f"执行图片描述插入时出错: {e}")
        print("===== 图片描述插入执行失败 =====\n")
        return False

def get_bat_directory():
    """
    获取调用此脚本的bat文件所在目录
    """
    try:
        print("\n===== 开始获取BAT文件所在目录 =====")
        # 获取调用栈信息
        frame = inspect.stack()[1]
        module = inspect.getmodule(frame[0])
        print(f"当前模块文件: {module.__file__ if hasattr(module, '__file__') else '未知'}")
        print(f"sys.argv[0]: {sys.argv[0]}")
        
        # 如果是通过bat文件调用的，sys.argv[0]应该是脚本的完整路径
        if module.__file__ == sys.argv[0]:
            # 获取bat文件所在目录（即当前工作目录）
            bat_dir = os.getcwd()
            print(f"检测到通过BAT文件调用，目录为: {bat_dir}")
            print("===== BAT文件所在目录获取成功 =====\n")
            return bat_dir
        else:
            # 如果不是通过bat调用的，返回None
            print("未检测到通过BAT文件调用，返回None")
            print("===== BAT文件所在目录获取失败 =====\n")
            return None
    except Exception as e:
        print(f"获取bat文件目录时出错: {e}")
        print(f"错误类型: {type(e).__name__}")
        print("===== BAT文件所在目录获取失败 =====\n")
        return None

def read_yaml_config():
    """
    读取yaml配置文件中的项目路径
    
    Returns:
        项目路径，如果配置文件不存在或路径为空则返回None
    """
    try:
        print("\n===== 开始读取YAML配置文件 =====")
        # 获取yaml配置文件路径
        yaml_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                                "# 模型训练", "lora_training_config.yaml")
        print(f"YAML配置文件路径: {yaml_path}")
        
        # 检查配置文件是否存在
        if not os.path.exists(yaml_path):
            print(f"警告: 配置文件 {yaml_path} 不存在")
            print("===== YAML配置文件读取失败 =====\n")
            return None
        
        # 读取配置文件
        with open(yaml_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        
        # 检查配置文件中是否有项目路径
        if config and 'project_path' in config and config['project_path']:
            # 获取路径并标准化
            project_path = config['project_path']
            print(f"从配置文件中读取到原始项目路径: {project_path}")
            
            # 处理路径中的斜杠，确保使用操作系统兼容的路径格式
            # 将正斜杠或双反斜杠转换为系统标准路径格式
            project_path = os.path.normpath(project_path)
            print(f"标准化后的项目路径: {project_path}")
            
            # 检查路径是否存在
            if os.path.exists(project_path):
                print(f"项目路径验证成功: {project_path}")
                print("===== YAML配置文件读取成功 =====\n")
                return project_path
            else:
                print(f"警告: yaml配置文件中的项目路径 {project_path} 不存在")
                print(f"请检查路径是否正确，确保使用正斜杠(/)或双反斜杠(\\\\)表示路径")
        else:
            print("警告: yaml配置文件中未设置项目路径或项目路径为空")
        
        print("===== YAML配置文件读取失败 =====\n")
        return None
    except yaml.YAMLError as ye:
        print(f"解析yaml配置文件时出错: {ye}")
        print("请确保yaml格式正确，Windows路径中的反斜杠需要使用正斜杠(/)或双反斜杠(\\\\)表示")
        print("===== YAML配置文件读取失败 =====\n")
        return None
    except Exception as e:
        print(f"读取yaml配置文件时出错: {e}")
        print(f"错误类型: {type(e).__name__}")
        print("===== YAML配置文件读取失败 =====\n")
        return None

def is_direct_run():
    """
    检测是否是直接运行脚本
    
    Returns:
        如果是直接运行脚本返回True，否则返回False
    """
    # 获取调用栈信息
    frame = inspect.stack()[1]
    module = inspect.getmodule(frame[0])
    
    # 如果是通过命令行直接运行，__name__应该是"__main__"
    return __name__ == "__main__"

def shutdown_computer():
    """
    执行关机操作
    """
    try:
        print("\n===== 开始执行关机操作 =====")
        import os
        os.system("shutdown /s /t 0")
        print("===== 关机命令已发送 =====\n")
    except Exception as e:
        print(f"执行关机操作时出错: {e}")

def main():
    try:
        print("\n========================================")
        print("     Lora训练初始化脚本开始执行     ")
        print("========================================\n")
        # 解析命令行参数
        print("===== 开始解析命令行参数 =====")
        args = parse_arguments()
        print(f"解析到的参数: {args}")
        print("===== 命令行参数解析完成 =====\n")
        
        # 确定项目路径
        project_path = None
        
        # 检查是否是直接运行脚本
        direct_run = is_direct_run()
        
        # 如果是直接运行脚本或指定了使用yaml配置文件
        if direct_run or args.use_yaml:
            # 尝试从yaml配置文件中读取项目路径
            yaml_project_path = read_yaml_config()
            if yaml_project_path:
                project_path = yaml_project_path
                print(f"使用yaml配置文件中的项目路径: {project_path}")
        
        # 如果指定了使用bat文件所在目录且未从yaml中获取路径
        if args.bat_dir and not project_path:
            bat_dir = get_bat_directory()
            if bat_dir:
                project_path = bat_dir
                print(f"使用bat文件所在目录作为项目路径: {project_path}")
            else:
                print("无法获取bat文件所在目录，将使用其他方式确定项目路径")
        
        # 如果指定了项目路径参数且未从其他方式获取路径
        if args.project_path and not project_path:
            project_path = args.project_path
            print(f"使用命令行参数指定的项目路径: {project_path}")
        
        # 检测项目类型和名称
        project_type, project_name, project_path = detect_project_type_and_name(project_path)
        print(f"检测到项目类型: {project_type}, 项目名称: {project_name}")
        print(f"项目路径: {project_path}")
        
        # 检查训练信息Excel文件是否已存在
        print("\n===== 检查训练信息Excel文件 =====")
        excel_path = os.path.join(project_path, "训练信息.xlsx")
        print(f"Excel文件路径: {excel_path}")
        excel_exists = os.path.exists(excel_path)
        
        if excel_exists:
            print(f"检测到训练信息Excel文件已存在: {excel_path}")
            print("跳过初始化步骤，直接读取执行标志")
            print("===== 训练信息Excel文件检查完成 =====\n")
        else:
            print("训练信息Excel文件不存在，需要创建")
            print("===== 训练信息Excel文件检查完成 =====\n")
            
            # 创建目录结构
            create_directory_structure(project_path)
            
            # 创建训练信息Excel文件
            excel_path = create_training_info_excel(project_type, project_name, project_path)
            
            print("\n===== 项目初始化完成! =====")
            print(f"请在 {excel_path} 中设置训练步骤的执行标志")
            print("===== 初始化阶段结束 =====\n")
        
        # 读取执行标志
        execution_flags = read_execution_flags(excel_path)
        
        print("\n===== 开始执行训练流程 =====")
        print(f"共检测到 {len(execution_flags)} 个步骤")
        
        # 检查是否需要执行图片尺寸标准化
        step_name = "图片尺寸标准化"
        print(f"\n----- 步骤1: {step_name} -----")
        if execution_flags.get(step_name) == 1 and not check_step_completed(excel_path, step_name):
            print(f"执行标志为1，且步骤未完成，开始执行{step_name}...")
            resize_dir = os.path.join(project_path, "resize")
            print(f"resize目录: {resize_dir}")
            success = run_resize_script(resize_dir)
            # 更新Excel中的完成结果
            update_step_result(excel_path, step_name, success)
            print(f"----- 步骤1: {step_name} {'完成' if success else '失败'} -----")
        else:
            skip_reason = "步骤已完成" if check_step_completed(excel_path, step_name) else "执行标志未设置为1"
            print(f"跳过{step_name}步骤，原因: {skip_reason}")
            print(f"----- 步骤1: {step_name} 已跳过 -----")
        
        # 检查是否需要执行图片描述生成
        step_name = "图片描述生成"
        print(f"\n----- 步骤2: {step_name} -----")
        # 只有当前一步骤成功完成时，才执行当前步骤
        if execution_flags.get(step_name) == 1 and check_step_completed(excel_path, "图片尺寸标准化") and not check_step_completed(excel_path, step_name):
            print(f"执行标志为1，前置步骤已完成，且当前步骤未完成，开始执行{step_name}...")
            resize_dir = os.path.join(project_path, "resize")
            print(f"resize目录: {resize_dir}")
            success = run_image_description_script(resize_dir)
            # 更新Excel中的完成结果
            update_step_result(excel_path, step_name, success)
            print(f"----- 步骤2: {step_name} {'完成' if success else '失败'} -----")
        else:
            skip_reason = "步骤已完成" if check_step_completed(excel_path, step_name) else \
                        "前置步骤未完成" if not check_step_completed(excel_path, "图片尺寸标准化") else \
                        "执行标志未设置为1"
            print(f"跳过{step_name}步骤，原因: {skip_reason}")
            print(f"----- 步骤2: {step_name} 已跳过 -----")
        
        # 检查是否需要执行图片描述优化
        step_name = "图片描述优化"
        print(f"\n----- 步骤3: {step_name} -----")
        # 只有当前一步骤成功完成时，才执行当前步骤
        if execution_flags.get(step_name) == 1 and check_step_completed(excel_path, "图片描述生成") and not check_step_completed(excel_path, step_name):
            print(f"执行标志为1，前置步骤已完成，且当前步骤未完成，开始执行{step_name}...")
            success = run_description_optimization_script(project_path)
            # 更新Excel中的完成结果
            update_step_result(excel_path, step_name, success)
            print(f"----- 步骤3: {step_name} {'完成' if success else '失败'} -----")
        elif execution_flags.get(step_name) == 2 and check_step_completed(excel_path, "图片描述生成") and not check_step_completed(excel_path, step_name):
            print(f"执行标志为2，前置步骤已完成，且当前步骤未完成，开始执行{step_name}(含AI翻译)...")
            success = run_description_optimization_script(project_path, True)
            # 更新Excel中的完成结果
            update_step_result(excel_path, step_name, success)
            print(f"----- 步骤3: {step_name}(含AI翻译) {'完成' if success else '失败'} -----")
        else:
            skip_reason = "步骤已完成" if check_step_completed(excel_path, step_name) else \
                        "前置步骤未完成" if not check_step_completed(excel_path, "图片描述生成") else \
                        "执行标志未设置为1或2"
            print(f"跳过{step_name}步骤，原因: {skip_reason}")
            print(f"----- 步骤3: {step_name} 已跳过 -----")
        
        # 检查是否需要执行图片描述插入
        step_name = "图片描述插入"
        print(f"\n----- 步骤3.5: {step_name} -----")
        # 只有当前一步骤成功完成时，才执行当前步骤
        if execution_flags.get(step_name) == 1 and check_step_completed(excel_path, "图片描述优化") and not check_step_completed(excel_path, step_name):
            print(f"执行标志为1，前置步骤已完成，且当前步骤未完成，开始执行{step_name}...")
            success = run_description_insertion_script(project_path, excel_path)
            # 更新Excel中的完成结果
            update_step_result(excel_path, step_name, success)
            print(f"----- 步骤3.5: {step_name} {'完成' if success else '失败'} -----")
        else:
            skip_reason = "步骤已完成" if check_step_completed(excel_path, step_name) else \
                        "前置步骤未完成" if not check_step_completed(excel_path, "图片描述优化") else \
                        "执行标志未设置为1"
            print(f"跳过{step_name}步骤，原因: {skip_reason}")
            print(f"----- 步骤3.5: {step_name} 已跳过 -----")
        
        # 检查是否需要执行模型Lora训练
        step_name = "模型Lora训练"
        print(f"\n----- 步骤4: {step_name} -----")
        # 只有当前一步骤成功完成时，才执行当前步骤
        if execution_flags.get(step_name) == 1 and check_step_completed(excel_path, "图片描述优化") and not check_step_completed(excel_path, step_name):
            print(f"执行标志为1，前置步骤已完成，且当前步骤未完成，开始执行{step_name}...")
            success = run_model_training_script(project_path)
            # 更新Excel中的完成结果
            update_step_result(excel_path, step_name, success)
            print(f"----- 步骤4: {step_name} {'完成' if success else '失败'} -----")
        else:
            skip_reason = "步骤已完成" if check_step_completed(excel_path, step_name) else \
                        "前置步骤未完成" if not check_step_completed(excel_path, "图片描述优化") else \
                        "执行标志未设置为1"
            print(f"跳过{step_name}步骤，原因: {skip_reason}")
            print(f"----- 步骤4: {step_name} 已跳过 -----")
        
        # 检查是否需要执行模型Lora测试
        step_name = "模型Lora测试"
        print(f"\n----- 步骤5: {step_name} -----")
        # 只有当前一步骤成功完成时，才执行当前步骤
        if execution_flags.get(step_name) == 1 and check_step_completed(excel_path, "模型Lora训练") and not check_step_completed(excel_path, step_name):
            print(f"执行标志为1，前置步骤已完成，且当前步骤未完成，开始执行{step_name}...")
            success = run_model_test_script(project_path)
            # 更新Excel中的完成结果
            update_step_result(excel_path, step_name, success)
            print(f"----- 步骤5: {step_name} {'完成' if success else '失败'} -----")
        else:
            skip_reason = "步骤已完成" if check_step_completed(excel_path, step_name) else \
                        "前置步骤未完成" if not check_step_completed(excel_path, "模型Lora训练") else \
                        "执行标志未设置为1"
            print(f"跳过{step_name}步骤，原因: {skip_reason}")
            print(f"----- 步骤5: {step_name} 已跳过 -----")
        
        print("\n===== 训练流程执行完成 =====")
        # 统计已完成的步骤
        completed_steps = [step for step in execution_flags.keys() 
                        if check_step_completed(excel_path, step)]
        print(f"已完成的步骤数: {len(completed_steps)}/{len(execution_flags)}")
        if completed_steps:
            print("已完成的步骤:")
            for step in completed_steps:
                print(f"  - {step}")
        print("===== 流程执行统计完成 =====\n")
        
        # 检查是否需要执行关机操作
        print("\n===== 检查是否需要执行关机操作 =====")
        shutdown_flag = execution_flags.get("是否关机", 0)
        print(f"是否关机标志: {shutdown_flag}")
        
        if shutdown_flag == 1 and check_step_completed(excel_path, "模型Lora测试"):
            print("是否关机标志为1，且模型Lora测试步骤已成功完成，将执行关机操作")
            shutdown_computer()
        elif shutdown_flag == 2:
            print("是否关机标志为2，无论流程成功或失败，将执行关机操作")
            shutdown_computer()
        else:
            print("不满足关机条件，跳过关机操作")
        print("===== 关机检查完成 =====\n")
        
    except Exception as e:
        print(f"\n===== 执行过程中发生错误 =====")
        print(f"错误信息: {e}")
        print(f"错误类型: {type(e).__name__}")
        print("===== 脚本执行失败 =====\n")
        return 1
    
    print("\n========================================")
    print("     Lora训练初始化脚本执行完成     ")
    print("========================================\n")
    return 0

if __name__ == "__main__":
    sys.exit(main())