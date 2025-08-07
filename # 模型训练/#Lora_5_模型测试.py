import os
import sys
import time
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
import io
import traceback

# 导入自定义工具包
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# 定义训练集类型和地址
CHARACTER_DIR = "E:\\Design\\Character"
STYLES_DIR = "E:\\Design\\Styles"
PROMPT_TEMPLATE_PATH = "E:\\Design\\提示词模板.xlsx"

class LoraModelTester:
    def __init__(self, project_path):
        self.project_path = project_path
        self.project_type = None
        self.project_name = None
        self.excel_path = None
        self.prompt_id = None
        self.trigger_word = None
        self.test_excel_path = None
        self.prompts = []
        self.version = None
        
    def detect_project_info(self):
        """检测项目类型和名称"""
        if self.project_path.startswith(CHARACTER_DIR):
            self.project_type = "Character"
            self.project_name = os.path.basename(self.project_path)
        elif self.project_path.startswith(STYLES_DIR):
            self.project_type = "Styles"
            self.project_name = os.path.basename(self.project_path)
        else:
            raise ValueError(f"项目路径 {self.project_path} 不在有效的训练集目录下")
        
        print(f"项目类型: {self.project_type}, 项目名称: {self.project_name}")
        
        # 查找训练信息Excel文件
        self.excel_path = os.path.join(self.project_path, "训练信息.xlsx")
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"未找到训练信息Excel文件: {self.excel_path}")
    
    def read_prompt_info(self):
        """从训练信息Excel文件中读取提示词编号、触发词和版本号"""
        try:
            wb = load_workbook(self.excel_path)
            steps_sheet = wb["步骤"]
            
            # 查找提示词编号、触发词和版本号
            for row in range(2, steps_sheet.max_row + 1):
                step_name = steps_sheet.cell(row=row, column=1).value
                if step_name == "提示词编号":
                    self.prompt_id = steps_sheet.cell(row=row, column=2).value
                elif step_name == "触发词":
                    self.trigger_word = steps_sheet.cell(row=row, column=2).value
                elif step_name == "版本号":
                    self.version = steps_sheet.cell(row=row, column=2).value
            
            if not self.prompt_id:
                raise ValueError("未在Excel文件中找到提示词编号")
            
            # 如果没有找到版本号，设置默认值
            if not self.version:
                self.version = "v1"
                print("警告: 未在Excel文件中找到版本号，使用默认值 v1")
            
            print(f"提示词编号: {self.prompt_id}, 触发词: {self.trigger_word}, 版本号: {self.version}")
            return True
        except Exception as e:
            print(f"读取提示词信息时出错: {str(e)}")
            traceback.print_exc()
            return False
    
    def read_prompt_template(self):
        """从提示词模板Excel文件中读取提示词列表"""
        try:
            if not os.path.exists(PROMPT_TEMPLATE_PATH):
                raise FileNotFoundError(f"未找到提示词模板文件: {PROMPT_TEMPLATE_PATH}")
            
            wb = load_workbook(PROMPT_TEMPLATE_PATH)
            
            # 检查是否存在对应的工作簿
            if self.prompt_id not in wb.sheetnames:
                raise ValueError(f"提示词模板中不存在工作簿: {self.prompt_id}")
            
            prompt_sheet = wb[self.prompt_id]
            
            # 读取提示词列表
            self.prompts = []
            for row in range(1, 11):  # 最多读取10个提示词
                prompt = prompt_sheet.cell(row=row, column=1).value
                if prompt:
                    self.prompts.append(prompt)
                else:
                    break
            
            if not self.prompts:
                raise ValueError(f"提示词模板中未找到有效的提示词")
            
            print(f"读取到 {len(self.prompts)} 个提示词")
            return True
        except Exception as e:
            print(f"读取提示词模板时出错: {str(e)}")
            traceback.print_exc()
            return False
    
    def create_test_excel(self):
        """创建测试Excel文件"""
        try:
            # 创建gemini目录（如果不存在）
            gemini_dir = os.path.join(self.project_path, "gemini")
            os.makedirs(gemini_dir, exist_ok=True)
            
            # 创建测试Excel文件
            self.test_excel_path = os.path.join(gemini_dir, "测试.xlsx")
            wb = Workbook()
            
            # 创建参数工作簿
            params_sheet = wb.active
            params_sheet.title = "参数"
            
            # 设置列宽
            params_sheet.column_dimensions['A'].width = 20
            params_sheet.column_dimensions['B'].width = 40
            params_sheet.column_dimensions['C'].width = 20
            params_sheet.column_dimensions['D'].width = 20
            
            # 添加标题行
            params_sheet['A1'] = "名称"
            params_sheet['B1'] = "值"
            params_sheet['C1'] = "节点名称"
            params_sheet['D1'] = "节点属性"
            
            # 设置标题行样式
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            header_font = Font(bold=True)
            for cell in params_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加固定参数
            fixed_params = [
                {"名称": "workflow", "值": "#5 提示词单次测试-Unet.json", "节点名称": "", "节点属性": ""},
                {"名称": "默认底模", "值": "FLUX\\flux1-dev-fp8.safetensors", "节点名称": "UNet加载器", "节点属性": "unet_name"},
                {"名称": "laten_x", "值": 1024, "节点名称": "空Latent图像", "节点属性": "width"},
                {"名称": "laten_y", "值": 1024, "节点名称": "空Latent图像", "节点属性": "height"},
                {"名称": "seed", "值": 441951251478752, "节点名称": "K采样器", "节点属性": "seed"},
                {"名称": "保存图片路径", "值": "F:\\TEST", "节点名称": "", "节点属性": ""},
            ]
            
            row = 2
            for param in fixed_params:
                params_sheet.cell(row=row, column=1, value=param["名称"])
                params_sheet.cell(row=row, column=2, value=param["值"])
                params_sheet.cell(row=row, column=3, value=param["节点名称"])
                params_sheet.cell(row=row, column=4, value=param["节点属性"])
                row += 1
            
            # 添加变动参数
            params_sheet.cell(row=row, column=1, value="提示词编号")
            params_sheet.cell(row=row, column=2, value=self.prompt_id)
            row += 1
            
            params_sheet.cell(row=row, column=1, value="提示词数量")
            params_sheet.cell(row=row, column=2, value=len(self.prompts))
            params_sheet.cell(row=row, column=3, value="提示词数量")
            params_sheet.cell(row=row, column=4, value="int")
            row += 1
            
            # 添加提示词
            for i, prompt in enumerate(self.prompts, 1):
                params_sheet.cell(row=row, column=1, value=f"正面-{i}")
                params_sheet.cell(row=row, column=2, value=prompt)
                params_sheet.cell(row=row, column=3, value=f"String-{i}")
                params_sheet.cell(row=row, column=4, value="string")
                row += 1
            
            # 创建Lora-1工作簿
            lora_sheet = wb.create_sheet(title="Lora-1")
            
            # 设置列宽
            lora_sheet.column_dimensions['A'].width = 40
            lora_sheet.column_dimensions['B'].width = 20
            lora_sheet.column_dimensions['C'].width = 15
            lora_sheet.column_dimensions['D'].width = 15
            lora_sheet.column_dimensions['E'].width = 20
            lora_sheet.column_dimensions['F'].width = 20
            lora_sheet.column_dimensions['G'].width = 40
            lora_sheet.column_dimensions['H'].width = 20
            
            # 添加标题行
            lora_sheet['A1'] = "值"
            lora_sheet['B1'] = "节点名称"
            lora_sheet['C1'] = "lora强度"
            lora_sheet['D1'] = "clip强度"
            lora_sheet['E1'] = "编号"
            lora_sheet['F1'] = "触发词"
            lora_sheet['G1'] = "图片路径"
            lora_sheet['H1'] = "展示图片"
            
            # 设置标题行样式
            for cell in lora_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 查找项目目录下的safetensors文件
            lora_dir = f"E:/Design/loras/{self.project_name}"
            if os.path.exists(lora_dir):
                row = 2
                # 按文件名排序，没有后缀的排最后
                lora_files = []
                for file in os.listdir(lora_dir):
                    if file.endswith(".safetensors"):
                        lora_files.append(file)
                
                # 排序文件，没有后缀的排最后
                def get_suffix(filename):
                    parts = filename.split("-")
                    if len(parts) > 1 and parts[-1].split(".")[0].isdigit():
                        return int(parts[-1].split(".")[0])
                    return float('inf')  # 没有后缀的排最后
                
                lora_files.sort(key=get_suffix)
                
                for file in lora_files:
                    file_path = os.path.join(lora_dir, file)
                    # 从路径中提取值（去掉E:/Design/loras/前缀）
                    value = file_path.replace("E:/Design/loras/", "")
                    
                    # 提取编号（项目名称-版本号-文件名后缀取最后三位数）
                    suffix = "last"
                    parts = file.split("-")
                    if len(parts) > 1 and parts[-1].split(".")[0].isdigit():
                        suffix = parts[-1].split(".")[0]
                    
                    # 将版本号插入到model_id的中间
                    model_id = f"{self.project_name}-{self.version}-{suffix}"
                    
                    lora_sheet.cell(row=row, column=1, value=value)
                    lora_sheet.cell(row=row, column=2, value="Load LoRA-1")
                    lora_sheet.cell(row=row, column=3, value=1)
                    lora_sheet.cell(row=row, column=4, value=1)
                    lora_sheet.cell(row=row, column=5, value=model_id)
                    lora_sheet.cell(row=row, column=6, value=self.trigger_word)
                    row += 1
            
            # 保存工作簿
            wb.save(self.test_excel_path)
            print(f"已创建测试Excel文件: {self.test_excel_path}")
            return True
        except Exception as e:
            print(f"创建测试Excel文件时出错: {str(e)}")
            traceback.print_exc()
            return False

    def check_test_completion(self):
        """检查测试是否真正完成并成功"""
        try:
            # 检查测试Excel文件是否存在
            if not os.path.exists(self.test_excel_path):
                print(f"测试Excel文件不存在: {self.test_excel_path}")
                return False
            
            # 获取E:\Design\loras\项目名称目录下的safetensors文件数量
            expected_models_count = 0
            lora_dir = f"E:/Design/loras/{self.project_name}"
            if os.path.exists(lora_dir):
                for file in os.listdir(lora_dir):
                    if file.endswith(".safetensors"):
                        expected_models_count += 1
            
            if expected_models_count == 0:
                print(f"警告: 在 {lora_dir} 目录下没有找到任何safetensors文件")
                return False
                
            # 重新加载Excel文件检查是否有图片路径
            wb = load_workbook(self.test_excel_path)
            success = False
            
            # 检查Lora-1工作簿中是否有图片路径
            if "Lora-1" in wb.sheetnames:
                lora_sheet = wb["Lora-1"]
                headers = [cell.value for cell in lora_sheet[1]]
                if "图片路径" in headers and "值" in headers:
                    img_path_idx = headers.index("图片路径")
                    value_idx = headers.index("值")
                    
                    # 检查是否有任何行包含有效的图片路径
                    valid_images_count = 0
                    total_models_count = 0
                    
                    for row in range(2, lora_sheet.max_row + 1):
                        # 检查第一列是否有值
                        value = lora_sheet.cell(row=row, column=value_idx+1).value
                        if value and value !="":  # 如果第一列有值
                            total_models_count += 1
                            # 检查该行的图片路径是否存在且有效
                            img_path = lora_sheet.cell(row=row, column=img_path_idx+1).value
                            if img_path and os.path.exists(img_path):
                                # 验证图片文件是否可以打开
                                try:
                                    with Image.open(img_path) as img:
                                        # 检查图片尺寸是否合理
                                        if img.width > 100 and img.height > 100:
                                            valid_images_count += 1
                                        else:
                                            print(f"行 {row} 的图片尺寸异常: {img.width}x{img.height}")
                                except Exception as e:
                                    print(f"行 {row} 的图片无法打开: {str(e)}")
                            else:
                                print(f"行 {row} 有值但没有有效的图片路径")
                    
                    # 验证Excel中的模型数量是否与目录中的safetensors文件数量一致
                    if total_models_count != expected_models_count:
                        print(f"警告: Excel中的模型数量({total_models_count})与目录中的safetensors文件数量({expected_models_count})不一致")
                    
                    # 只有当所有模型都成功生成了有效图片时，才认为测试成功
                    if total_models_count > 0 and valid_images_count == total_models_count:
                        success = True
                        print(f"测试成功: 所有 {total_models_count} 个模型都生成了有效图片")
                    else:
                        print(f"测试部分失败: 共 {total_models_count} 个模型，但只有 {valid_images_count} 个生成了有效图片")
            
            if not success:
                print("批量测试未能成功生成所有图片")
                return False
                
            print("批量测试已完成，成功生成了所有图片")
            return True
        except Exception as e:
            print(f"检查测试完成状态时出错: {str(e)}")
            traceback.print_exc()
            return False

    
    def update_excel_status(self, status):
        """更新训练信息Excel文件中的测试状态"""
        try:
            if not os.path.exists(self.excel_path):
                print(f"训练信息Excel文件不存在: {self.excel_path}")
                return False
                
            wb = load_workbook(self.excel_path)
            if "步骤" in wb.sheetnames:
                steps_sheet = wb["步骤"]
                
                # 查找测试状态行
                status_row = None
                for row in range(2, steps_sheet.max_row + 1):
                    step_name = steps_sheet.cell(row=row, column=1).value
                    if step_name == "测试状态":
                        status_row = row
                        break
                
                # 如果找到了测试状态行，更新状态
                if status_row:
                    steps_sheet.cell(row=status_row, column=2, value=status)
                else:
                    # 如果没有找到测试状态行，添加一行
                    steps_sheet.cell(row=steps_sheet.max_row + 1, column=1, value="测试状态")
                    steps_sheet.cell(row=steps_sheet.max_row, column=2, value=status)
                
                # 保存Excel文件
                wb.save(self.excel_path)
                print(f"已更新测试状态为: {status}")
                return True
            else:
                print(f"训练信息Excel文件中没有步骤工作簿")
                return False
        except Exception as e:
            print(f"更新测试状态时出错: {str(e)}")
            traceback.print_exc()
            return False
    
    def run(self):
        """运行测试流程"""
        try:
            # 检测项目信息
            self.detect_project_info()
            
            # 读取提示词信息
            if not self.read_prompt_info():
                self.update_excel_status("失败：无法读取提示词信息")
                return False
            
            # 读取提示词模板
            if not self.read_prompt_template():
                self.update_excel_status("失败：无法读取提示词模板")
                return False
            
            # 创建测试Excel文件
            if not self.create_test_excel():
                self.update_excel_status("失败：无法创建测试Excel文件")
                return False
            
            # 检查测试是否真正完成并成功
            test_completion_result = self.check_test_completion()
            if not test_completion_result:
                self.update_excel_status("失败：测试未能生成所有有效图片")
                return False
            else:
                # 更新状态为成功
                self.update_excel_status("成功")
                return True
        except Exception as e:
            print(f"测试流程执行出错: {str(e)}")
            traceback.print_exc()
            self.update_excel_status(f"失败：{str(e)}")
            return False


def main():
    # 解析命令行参数
    import argparse
    parser = argparse.ArgumentParser(description="Lora模型测试脚本")
    parser.add_argument("--project_path", type=str, help="项目路径")
    args = parser.parse_args()
    
    # 如果没有提供项目路径，使用当前工作目录
    project_path = args.project_path if args.project_path else os.getcwd()
    
    # 创建测试实例并运行
    tester = LoraModelTester(project_path)
    tester.run()


if __name__ == "__main__":
    main()