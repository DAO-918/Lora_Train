import os
import sys
import json
import time
import shutil
import openpyxl
from PIL import Image
import importlib.util
import traceback
from pathlib import Path

# 导入自定义工具包
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.comfy_websocket_wrapper import ComfyWebSocketClient
from utils.comfy_workflow_wrapper import ComfyWorkflowWrapper

# 尝试导入翻译模块
try:
    # 动态导入模块
    translate_baidu_spec = importlib.util.find_spec('utils.translate_baidu_request')
    translate_tencent_spec = importlib.util.find_spec('utils.translate_tencent_request')
    
    if translate_baidu_spec:
        translate_baidu = importlib.util.module_from_spec(translate_baidu_spec)
        translate_baidu_spec.loader.exec_module(translate_baidu)
    
    if translate_tencent_spec:
        translate_tencent = importlib.util.module_from_spec(translate_tencent_spec)
        translate_tencent_spec.loader.exec_module(translate_tencent)
    
    has_translate_modules = translate_baidu_spec is not None or translate_tencent_spec is not None
except ImportError:
    has_translate_modules = False
    print("警告: 翻译模块未找到，将使用Gemini进行翻译")


class DescriptionOptimizer:
    def __init__(self, project_dir, optimization_mode=1):
        """
        初始化描述优化器
        
        Args:
            project_dir: 项目根目录
            optimization_mode: 优化模式，1=英文优化，2=中文优化并翻译
        """
        self.project_dir = project_dir
        self.optimization_mode = optimization_mode
        self.gemini_folder_path = os.path.join(project_dir, "gemini")
        self.excel_path = os.path.join(project_dir, "训练信息.xlsx")
        self.updated_files = []  # 记录已更新的文件
        self.failed_files = []   # 记录失败的文件
        
        # 检查必要的文件和文件夹
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"错误: 未找到训练信息.xlsx文件: {self.excel_path}")
            
        if not os.path.exists(self.gemini_folder_path):
            raise FileNotFoundError(f"错误: 未找到gemini文件夹: {self.gemini_folder_path}")
    
    def translate_text(self, text, to_lang="en"):
        """
        翻译文本
        
        Args:
            text: 要翻译的文本
            to_lang: 目标语言，默认为英文
            
        Returns:
            翻译后的文本，如果失败则返回空字符串
        """
        print(f"      * 开始翻译文本，长度: {len(text)} 字符，目标语言: {to_lang}")
        
        if not has_translate_modules:
            # 如果没有翻译模块，使用Gemini进行翻译
            print(f"      * 未找到翻译模块，使用Gemini进行翻译")
            try:
                import google.generativeai as genai
                
                # 尝试从环境变量或配置文件获取API密钥
                print(f"      * 尝试获取Gemini API密钥...")
                api_key = os.environ.get("GEMINI_API_KEY", "")
                if not api_key:
                    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "gemini_config.json")
                    print(f"      * 尝试从配置文件获取API密钥: {config_path}")
                    if os.path.exists(config_path):
                        with open(config_path, "r") as f:
                            config = json.load(f)
                            api_key = config.get("GEMINI_API_KEY", "")
                
                if not api_key:
                    print("      * 错误: 未找到Gemini API密钥，无法进行翻译")
                    return ""
                else:
                    print("      * 成功获取API密钥")
                
                print("      * 配置Gemini API并创建模型...")
                genai.configure(api_key=api_key)
                model = genai.GenerativeModel('gemini-pro')
                
                # 根据目标语言设置提示词
                if to_lang.lower() == "en":
                    prompt = f"Translate the following Chinese text to English. Only provide the translation without any additional text or explanation:\n\n{text}"
                else:
                    prompt = f"Translate the following English text to Chinese. Only provide the translation without any additional text or explanation:\n\n{text}"
                
                print("      * 发送翻译请求到Gemini API...")
                response = model.generate_content(prompt)
                
                result = response.text.strip()
                print(f"      * 翻译成功，结果长度: {len(result)} 字符")
                return result
            except Exception as e:
                print(f"      * 错误: 使用Gemini翻译时出现异常: {str(e)}")
                traceback.print_exc()
                return ""
        else:
            # 使用翻译模块
            print(f"      * 使用翻译模块进行翻译")
            try:
                # 先尝试百度翻译
                if translate_baidu_spec:
                    print(f"      * 尝试使用百度翻译...")
                    try:
                        result = translate_baidu.translate_text(text, to_lang)
                        if result:
                            print(f"      * 百度翻译成功，结果长度: {len(result)} 字符")
                            return result
                        else:
                            print(f"      * 百度翻译返回空结果")
                    except Exception as e:
                        print(f"      * 警告: 百度翻译失败，尝试腾讯翻译: {str(e)}")
                
                # 如果百度翻译失败，尝试腾讯翻译
                if translate_tencent_spec:
                    print(f"      * 尝试使用腾讯翻译...")
                    try:
                        result = translate_tencent.translate_text(text, to_lang)
                        if result:
                            print(f"      * 腾讯翻译成功，结果长度: {len(result)} 字符")
                            return result
                        else:
                            print(f"      * 腾讯翻译返回空结果")
                    except Exception as e:
                        print(f"      * 警告: 腾讯翻译失败: {str(e)}")
                
                # 如果两个都失败，返回空字符串
                print(f"      * 所有翻译方法都失败")
                return ""
            except Exception as e:
                print(f"      * 错误: 翻译时出现异常: {str(e)}")
                traceback.print_exc()
                return ""
    
    def process_descriptions(self):
        """
        处理所有图片描述
        """
        print("\n===== 开始处理图片描述优化任务 =====")
        print(f"项目目录: {self.project_dir}")
        print(f"优化模式: {self.optimization_mode} ({'英文优化' if self.optimization_mode == 1 else '中文优化并翻译'})")
        print(f"Excel文件: {self.excel_path}")
        print(f"Gemini文件夹: {self.gemini_folder_path}")
        
        try:
            # 打开Excel文件
            print("\n[1/4] 正在读取Excel文件...")
            wb = openpyxl.load_workbook(self.excel_path)
            print(f"成功打开Excel文件: {self.excel_path}")
            
            # 检查是否有"提示词"工作表
            if "提示词" not in wb.sheetnames:
                print("错误: Excel文件中没有'提示词'工作表")
                return False
            
            ws = wb["提示词"]
            print(f"成功找到'提示词'工作表")
            
            # 获取所有图片路径和提示词
            print("\n[2/4] 正在收集图片数据...")
            image_data = []
            for row in range(1, ws.max_row + 1):
                english_prompt = ws.cell(row=row, column=1).value
                chinese_prompt = ws.cell(row=row, column=2).value
                image_path = ws.cell(row=row, column=4).value
                
                if image_path and (english_prompt or chinese_prompt):
                    image_data.append({
                        "row": row,
                        "english_prompt": english_prompt,
                        "chinese_prompt": chinese_prompt,
                        "image_path": image_path
                    })
            
            if not image_data:
                print("警告: 未找到任何图片数据")
                return False
            
            print(f"成功收集 {len(image_data)} 条图片数据")
            
            # 处理每个图片的描述
            print("\n[3/4] 开始处理图片描述...")
            total_count = len(image_data)
            processed_count = 0
            skipped_count = 0
            
            for data in image_data:
                processed_count += 1
                image_name = os.path.basename(data["image_path"])
                print(f"\n处理图片 [{processed_count}/{total_count}]: {image_name}")
                result = self.process_single_description(data)
                if result == "skipped":
                    skipped_count += 1
            
            # 保存Excel文件
            print("\n[4/4] 正在保存Excel文件...")
            wb.save(self.excel_path)
            print(f"Excel文件已保存: {self.excel_path}")
            
            # 更新步骤工作表的完成结果
            self.update_step_status()
            
            # 输出统计信息
            print("\n===== 图片描述优化任务完成 =====")
            print(f"总计处理: {total_count} 个文件")
            print(f"已更新: {len(self.updated_files)} 个文件")
            print(f"已跳过: {skipped_count} 个文件")
            print(f"失败: {len(self.failed_files)} 个文件")
            
            return True
            
        except Exception as e:
            print(f"错误: 处理描述时出现异常: {str(e)}")
            traceback.print_exc()
            return False
    
    def process_single_description(self, data):
        """
        处理单个图片描述
        
        Args:
            data: 包含图片数据的字典
            
        Returns:
            处理结果: "updated", "skipped", "failed"
        """
        try:
            # 获取图片对应的txt文件路径
            image_path = data["image_path"]
            image_name = os.path.basename(image_path)
            image_dir = os.path.dirname(image_path)
            
            # 检查图片是否在gemini文件夹中
            if not image_dir.startswith(self.gemini_folder_path):
                print(f"  - 警告: 图片不在gemini文件夹中: {image_path}")
                return "skipped"
            
            # 获取txt文件路径
            print(f"  - 查找对应的txt文件...")
            txt_path = os.path.splitext(image_path)[0] + ".txt"
            if not os.path.exists(txt_path):
                print(f"  - 未找到正常txt文件，尝试查找错误文件...")
                txt_path = os.path.splitext(image_path)[0] + ".error.txt"  # 尝试查找错误文件
                if not os.path.exists(txt_path):
                    print(f"  - 警告: 未找到任何对应的txt文件: {txt_path}")
                    return "skipped"
                else:
                    print(f"  - 找到错误txt文件: {os.path.basename(txt_path)}")
            else:
                print(f"  - 找到txt文件: {os.path.basename(txt_path)}")
            
            # 读取当前txt文件内容
            print(f"  - 读取当前txt文件内容...")
            with open(txt_path, "r", encoding="utf-8") as f:
                current_txt_content = f.read().strip()
            print(f"  - 当前内容长度: {len(current_txt_content)} 字符")
            
            # 根据优化模式处理
            if self.optimization_mode == 1:  # 英文优化
                print(f"  - 使用英文优化模式处理...")
                # 检查英文内容是否发生变化
                if not data["english_prompt"]:
                    print(f"  - 警告: Excel中英文提示词为空，跳过处理")
                    return "skipped"
                    
                if data["english_prompt"] == current_txt_content:
                    print(f"  - 内容未变化，无需更新")
                    return "skipped"
                else:
                    print(f"  - 检测到内容变化，更新txt文件...")
                    # 更新txt文件
                    with open(txt_path, "w", encoding="utf-8") as f:
                        f.write(data["english_prompt"])
                    print(f"  - 已更新英文描述: {os.path.basename(txt_path)}")
                    self.updated_files.append(os.path.basename(txt_path))
                    return "updated"
            
            elif self.optimization_mode == 2:  # 中文优化并翻译
                print(f"  - 使用中文优化并翻译模式处理...")
                # 检查中文内容是否发生变化
                if not data["chinese_prompt"]:
                    print(f"  - 警告: Excel中中文提示词为空，跳过处理")
                    return "skipped"
                
                # 翻译成英文
                print(f"  - 开始翻译中文提示词...")
                english_prompt = self.translate_text(data["chinese_prompt"], "en")
                if english_prompt:
                    print(f"  - 翻译成功，长度: {len(english_prompt)} 字符")
                    # 更新txt文件
                    with open(txt_path, "w", encoding="utf-8") as f:
                        f.write(english_prompt)
                    print(f"  - 已更新英文描述(从中文翻译): {os.path.basename(txt_path)}")
                    self.updated_files.append(os.path.basename(txt_path))
                    
                    # 更新Excel中的英文提示词
                    print(f"  - 更新Excel中的英文提示词...")
                    wb = openpyxl.load_workbook(self.excel_path)
                    ws = wb["提示词"]
                    ws.cell(row=data["row"], column=1, value=english_prompt)
                    wb.save(self.excel_path)
                    print(f"  - Excel文件已更新")
                    return "updated"
                else:
                    print(f"  - 警告: 翻译失败: {os.path.basename(txt_path)}")
                    self.failed_files.append(os.path.basename(txt_path))
                    return "failed"
            
            return "skipped"
            
        except Exception as e:
            print(f"  - 错误: 处理单个描述时出现异常: {str(e)}")
            traceback.print_exc()
            self.failed_files.append(os.path.basename(image_path))
            return "failed"
    
    def update_step_status(self):
        """
        更新步骤工作表的完成结果
        """
        print("\n[更新Excel] 正在更新步骤工作表的完成结果...")
        try:
            # 打开Excel文件
            print("[更新Excel] 打开Excel文件...")
            wb = openpyxl.load_workbook(self.excel_path)
            
            # 检查是否有"步骤"工作表
            if "步骤" not in wb.sheetnames:
                print("[更新Excel] 错误: Excel文件中没有'步骤'工作表")
                return
            
            steps_ws = wb["步骤"]
            print("[更新Excel] 成功找到'步骤'工作表")
            
            # 查找图片描述优化步骤
            step_found = False
            for r in range(1, 10):  # 假设步骤不超过10行
                step_name = steps_ws.cell(row=r, column=1).value
                if step_name == "图片描述优化":
                    step_found = True
                    print(f"[更新Excel] 找到'图片描述优化'步骤，位于第{r}行")
                    
                    # 更新时间
                    current_time = time.strftime("%Y-%m-%d %H:%M:%S")
                    steps_ws.cell(row=r, column=3, value=current_time)
                    print(f"[更新Excel] 更新完成时间: {current_time}")
                    
                    # 更新状态
                    status = ""
                    if not self.failed_files:  # 如果没有失败的文件
                        if self.updated_files:  # 如果有更新的文件
                            status = "成功"
                            steps_ws.cell(row=r, column=4, value=status)
                        else:  # 如果没有更新的文件
                            status = "无需更新"
                            steps_ws.cell(row=r, column=4, value=status)
                    else:  # 如果有失败的文件
                        # 格式化失败信息
                        failed_info = "，".join([f"{name}：失败" for name in self.failed_files])
                        status = failed_info
                        steps_ws.cell(row=r, column=4, value=failed_info)
                    
                    print(f"[更新Excel] 更新状态: {status}")
                    break
            
            if not step_found:
                print("[更新Excel] 警告: 未找到'图片描述优化'步骤")
            
            # 保存Excel文件
            print("[更新Excel] 保存Excel文件...")
            wb.save(self.excel_path)
            print("[更新Excel] Excel文件已保存")
            
        except Exception as e:
            print(f"[更新Excel] 错误: 更新步骤状态时出现异常: {str(e)}")
            traceback.print_exc()


def main():
    print("\n========================================")
    print("   Lora模型训练 - 图片描述优化工具   ")
    print("========================================\n")
    
    start_time = time.time()
    
    # 检查命令行参数
    print("[初始化] 检查命令行参数...")
    if len(sys.argv) < 2:
        print("用法: python #Lora_3_画面描述优化-Gemini.py <项目根目录> [优化模式]")
        print("优化模式: 1=英文优化(默认), 2=中文优化并翻译")
        return
    
    project_dir = sys.argv[1]
    print(f"[初始化] 项目目录: {project_dir}")
    
    if not os.path.exists(project_dir):
        print(f"[初始化] 错误: 路径不存在: {project_dir}")
        return
    else:
        print(f"[初始化] 项目目录存在，继续处理")
    
    # 获取优化模式
    optimization_mode = 1  # 默认为英文优化
    if len(sys.argv) > 2:
        try:
            optimization_mode = int(sys.argv[2])
            if optimization_mode not in [1, 2]:
                print(f"[初始化] 警告: 无效的优化模式: {optimization_mode}，使用默认值1")
                optimization_mode = 1
            else:
                print(f"[初始化] 优化模式: {optimization_mode} ({'英文优化' if optimization_mode == 1 else '中文优化并翻译'})")
        except ValueError:
            print(f"[初始化] 警告: 无效的优化模式: {sys.argv[2]}，使用默认值1")
            optimization_mode = 1
    else:
        print(f"[初始化] 使用默认优化模式: 1 (英文优化)")
    
    # 创建描述优化器并处理描述
    try:
        print("\n[初始化] 创建描述优化器...")
        optimizer = DescriptionOptimizer(project_dir, optimization_mode)
        print("[初始化] 描述优化器创建成功，开始处理描述")
        
        success = optimizer.process_descriptions()
        
        # 计算执行时间
        end_time = time.time()
        execution_time = end_time - start_time
        minutes, seconds = divmod(execution_time, 60)
        
        print("\n========================================")
        if success:
            print(f"✅ 描述优化任务完成")
            print(f"- 优化模式: {optimization_mode} ({'英文优化' if optimization_mode == 1 else '中文优化并翻译'})")
            print(f"- 执行时间: {int(minutes)}分{int(seconds)}秒")
            if optimizer.updated_files:
                print(f"- 已更新: {len(optimizer.updated_files)} 个文件")
                if len(optimizer.updated_files) <= 10:  # 如果更新的文件不多，显示文件名
                    for file in optimizer.updated_files:
                        print(f"  - {file}")
            else:
                print("- 没有文件需要更新")
            
            if optimizer.failed_files:
                print(f"- 失败: {len(optimizer.failed_files)} 个文件")
                if len(optimizer.failed_files) <= 10:  # 如果失败的文件不多，显示文件名
                    for file in optimizer.failed_files:
                        print(f"  - {file}")
        else:
            print("❌ 描述优化任务失败")
            print(f"- 执行时间: {int(minutes)}分{int(seconds)}秒")
        print("========================================\n")
    
    except Exception as e:
        print("\n========================================")
        print(f"❌ 错误: {str(e)}")
        traceback.print_exc()
        print("========================================\n")


if __name__ == "__main__":
    main()