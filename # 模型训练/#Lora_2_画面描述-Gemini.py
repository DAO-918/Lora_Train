import os
import sys
import json
import time
import shutil
import openpyxl
from PIL import Image
import importlib.util
import traceback
from pathlib import Path
from io import BytesIO

# 导入自定义工具包
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.comfy_websocket_wrapper import ComfyWebSocketClient
from utils.comfy_workflow_wrapper import ComfyWorkflowWrapper

# 尝试导入翻译模块
try:
    # 动态导入模块
    translate_baidu_spec = importlib.util.find_spec('utils.translate_baidu_request')
    translate_tencent_spec = importlib.util.find_spec('utils.translate_tencent_request')
    
    if translate_baidu_spec:
        translate_baidu = importlib.util.module_from_spec(translate_baidu_spec)
        translate_baidu_spec.loader.exec_module(translate_baidu)
    
    if translate_tencent_spec:
        translate_tencent = importlib.util.module_from_spec(translate_tencent_spec)
        translate_tencent_spec.loader.exec_module(translate_tencent)
    
    has_translate_modules = translate_baidu_spec is not None or translate_tencent_spec is not None
except ImportError:
    has_translate_modules = False
    print("警告: 翻译模块未找到，无法进行翻译")


class ImageDescriptionGenerator:
    def __init__(self, resize_folder_path):
        self.resize_folder_path = resize_folder_path
        self.gemini_folder_path = None
        self.server_address = "127.0.0.1:8191"
        self.client = ComfyWebSocketClient(self.server_address)
        self.workflow_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                                        "workflow", "#9 单次图片提示词生成.json")
        self.workflow = ComfyWorkflowWrapper(self.workflow_path)
        self.ai_prompt = """
Try to describe the content of the picture in detail, capturing every element without omission. Answer the description directly, without introductory phrases.
When there are people in the image, describe the appearance, clothes, movements and demeanor. Don't describe the names of the characters.
Please describe the content of this picture in detail, including the subject, background, color, style, etc. 
Specify the camera angle (high angle, low angle, close-up, mid-range, wide shot, etc.).
Indicate the depth of field (sharp focus, blurred background, deep focus, etc.).
Do NOT mention any text present in the image.
No ambiguous description is needed, brackets are not used to note the uncertain content, and the uncertain content can be omitted. Avoid ambiguous language—be precise and detailed.
"""
        self.error_strings = ["Error", "error", "failed"]
        self.excel_path = None
        self.current_image_path = None
        self.current_image_name = None
        self.picture_prompt = ""
        self.max_retries = 2
        self.failed_images = []  # 用于记录处理失败的图片名称
        self.reverse_prompt = ""  # 用于存储反推提示词
        
    def setup_gemini_folder(self):
        """设置gemini文件夹路径并创建文件夹"""
        # 获取resize文件夹的父目录
        parent_dir = os.path.dirname(self.resize_folder_path)
        # 创建gemini文件夹
        self.gemini_folder_path = os.path.join(parent_dir, "gemini")
        os.makedirs(self.gemini_folder_path, exist_ok=True)
        
    def find_excel_file(self):
        """查找训练信息.xlsx文件"""
        parent_dir = os.path.dirname(self.resize_folder_path)
        excel_path = os.path.join(parent_dir, "训练信息.xlsx")
        if os.path.exists(excel_path):
            self.excel_path = excel_path
            return True
        return False
        
    def read_reverse_prompt(self):
        """从Excel文件中读取反推提示词"""
        if not self.excel_path or not os.path.exists(self.excel_path):
            return False
            
        try:
            # 打开Excel文件
            wb = openpyxl.load_workbook(self.excel_path)
            
            # 检查是否有"步骤"工作表
            if "步骤" not in wb.sheetnames:
                print("警告: Excel文件中没有'步骤'工作表，无法读取反推提示词")
                return False
            
            steps_ws = wb["步骤"]
            
            # 查找反推提示词步骤
            for r in range(1, 20):  # 假设步骤不超过20行
                step_name = steps_ws.cell(row=r, column=1).value
                if step_name == "反推提示词":
                    # 读取反推提示词
                    reverse_prompt = steps_ws.cell(row=r, column=2).value
                    if reverse_prompt and isinstance(reverse_prompt, str) and reverse_prompt.strip():
                        self.reverse_prompt = reverse_prompt.strip()
                        print(f"已读取反推提示词: {self.reverse_prompt}")
                        return True
                    break
            
            return False
        except Exception as e:
            print(f"错误: 读取反推提示词时出现异常: {str(e)}")
            traceback.print_exc()
            return False
    
    def process_images(self):
        """处理resize文件夹中的所有图片"""
        print("\n===== 开始处理图片描述生成任务 =====")
        print(f"源图片文件夹: {self.resize_folder_path}")
        
        # 设置gemini文件夹
        self.setup_gemini_folder()
        print(f"目标Gemini文件夹: {self.gemini_folder_path}")
        
        # 查找Excel文件
        if not self.find_excel_file():
            print("错误: 未找到训练信息.xlsx文件")
            return
        print(f"已找到训练信息Excel文件: {self.excel_path}")
            
        # 读取反推提示词
        has_reverse_prompt = self.read_reverse_prompt()
        print(f"反推提示词状态: {'已读取' if has_reverse_prompt else '未找到或为空'}")
        
        # 获取所有图片文件
        image_files = []
        for root, _, files in os.walk(self.resize_folder_path):
            for file in files:
                if file.lower().endswith((".jpg", ".jpeg", ".png", ".bmp", ".webp")):
                    image_files.append(os.path.join(root, file))
        
        if not image_files:
            print(f"警告: 在{self.resize_folder_path}中未找到图片文件")
            return
        
        print(f"找到 {len(image_files)} 张图片需要处理")
        
        # 清空失败图片列表
        self.failed_images = []
        
        # 创建一个列表来存储已成功处理的图片信息，用于批量更新Excel
        successful_images = []
        # 创建一个列表来存储已存在的成功处理文件
        existing_processed_files = []
        
        # 处理每个图片
        processed_count = 0
        skipped_count = 0
        
        for image_path in image_files:
            processed_count += 1
            self.current_image_path = image_path
            self.current_image_name = os.path.basename(image_path)
            
            print(f"\n[{processed_count}/{len(image_files)}] 处理图片: {self.current_image_name}")
            
            # 创建对应的gemini子文件夹结构
            relative_path = os.path.relpath(os.path.dirname(image_path), self.resize_folder_path)
            target_dir = os.path.join(self.gemini_folder_path, relative_path) if relative_path != '.' else self.gemini_folder_path
            os.makedirs(target_dir, exist_ok=True)
            
            # 检查是否已经存在成功处理的文件
            target_image_path = os.path.join(target_dir, self.current_image_name)
            target_txt_path_normal = os.path.splitext(target_image_path)[0] + ".txt"
            target_txt_path_error = os.path.splitext(target_image_path)[0] + ".error.txt"
            
            # 如果已存在正常的txt文件，跳过处理
            if os.path.exists(target_image_path) and os.path.exists(target_txt_path_normal):
                print(f"  - 已存在成功处理的文件，跳过处理: {target_txt_path_normal}")
                # 将已存在的成功处理文件添加到列表中
                existing_processed_files.append({
                    "image_path": target_image_path,
                    "txt_path": target_txt_path_normal
                })
                skipped_count += 1
                continue
            
            # 如果存在错误的txt文件，删除它
            if os.path.exists(target_txt_path_error):
                print(f"  - 发现错误文件，删除: {target_txt_path_error}")
                try:
                    os.remove(target_txt_path_error)
                except Exception as e:
                    print(f"  - 删除错误文件失败: {str(e)}")
            
            # 生成图片描述
            print(f"  - 开始生成图片描述...")
            success = self.generate_image_description()
            
            # 如果失败，记录到失败列表中
            if not success:
                self.failed_images.append(self.current_image_name)
                print(f"  - 图片描述生成失败")
            else:
                print(f"  - 图片描述生成成功")
                # 将成功处理的图片信息添加到列表中，用于批量更新Excel
                successful_images.append({
                    "image_path": target_image_path,
                    "prompt": self.picture_prompt
                })
            
            # 复制图片和创建描述文件
            target_txt_path = os.path.splitext(target_image_path)[0] + (".error.txt" if not success else ".txt")
            
            # 复制图片
            print(f"  - 复制图片到: {target_image_path}")
            shutil.copy2(image_path, target_image_path)
            
            # 创建描述文件
            print(f"  - 创建描述文件: {target_txt_path}")
            with open(target_txt_path, "w", encoding="utf-8") as f:
                f.write(self.picture_prompt)
            
            print(f"  - 处理完成: {self.current_image_name} - {'成功' if success else '失败'}")
        
        # 批量更新Excel文件中的提示词和图片信息
        if successful_images:
            print(f"\n批量更新Excel文件中的提示词和图片信息...")
            self.batch_update_excel(successful_images)
        
        # 处理完所有图片后，更新步骤工作表的完成结果
        print("\n更新步骤工作表的完成结果...")
        self.update_step_status()
        
        # 输出处理统计信息
        print("\n===== 图片处理统计 =====")
        print(f"总图片数: {len(image_files)}")
        print(f"处理图片数: {processed_count - skipped_count}")
        print(f"跳过图片数: {skipped_count}")
        print(f"成功图片数: {processed_count - skipped_count - len(self.failed_images)}")
        print(f"失败图片数: {len(self.failed_images)}")
        if self.failed_images:
            print(f"失败图片列表: {', '.join(self.failed_images)}")
        print("===== 处理完成 =====")
    
    def generate_image_description(self):
        """生成图片描述"""
        retry_count = 0
        success = False
        
        while retry_count <= self.max_retries and not success:
            try:
                current_attempt = retry_count + 1
                max_attempts = self.max_retries + 1
                print(f"  - 尝试生成描述 (尝试 {current_attempt}/{max_attempts})")
                
                # 设置工作流参数
                print(f"    * 加载图片: {self.current_image_path}")
                self.workflow.set_node_param("Image Load", "image_path", self.current_image_path)
                
                # 根据是否有反推提示词决定使用哪个提示词
                prompt_to_use = self.reverse_prompt if self.reverse_prompt else self.ai_prompt
                prompt_type = "反推提示词" if self.reverse_prompt else "默认AI提示词"
                print(f"    * 使用{prompt_type}")
                self.workflow.set_node_param("提示词", "string", prompt_to_use)
                
                # 执行工作流
                print(f"    * 提交工作流到ComfyUI服务器")
                prompt = dict(self.workflow)
                self.client.queue_prompt(prompt)
                
                # 等待结果文件
                self.picture_prompt = ""
                temp_file_found = False
                max_wait_time = 180  # 最大等待时间（秒）
                start_time = time.time()
                print(f"    * 等待结果文件 (最长等待时间: {max_wait_time}秒)")
                
                check_count = 0
                while not temp_file_found and time.time() - start_time < max_wait_time:
                    # 检查临时文件
                    temp_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "# 模型训练")
                    for file in os.listdir(temp_dir):
                        if file.startswith("picture_prompt_temp") and file.endswith(".txt"):
                            temp_file_path = os.path.join(temp_dir, file)
                            print(f"    * 找到临时文件: {file}")
                            with open(temp_file_path, "r", encoding="utf-8") as f:
                                self.picture_prompt = f.read().strip()
                            
                            # 删除临时文件
                            print(f"    * 读取完成，删除临时文件")
                            os.remove(temp_file_path)
                            temp_file_found = True
                            break
                    
                    if not temp_file_found:
                        check_count += 1
                        if check_count % 6 == 0:  # 每30秒输出一次等待信息
                            elapsed_time = int(time.time() - start_time)
                            print(f"    * 等待结果中... (已等待 {elapsed_time} 秒)")
                        time.sleep(5)  # 等待5秒后再次检查
                
                if not temp_file_found:
                    print(f"    * 错误: 未找到临时文件，图片处理可能失败: {self.current_image_name}")
                    retry_count += 1
                    continue
                
                # 检查是否有错误
                has_error = any(error in self.picture_prompt for error in self.error_strings)
                if has_error:
                    print(f"    * 警告: 图片描述生成包含错误，重试中")
                    print(f"    * 错误内容: {self.picture_prompt}")
                    retry_count += 1
                    continue
                
                # 输出生成的描述（截断显示）
                display_prompt = self.picture_prompt[:100] + "..." if len(self.picture_prompt) > 100 else self.picture_prompt
                print(f"    * 成功生成描述: {display_prompt}")
                success = True
                
            except Exception as e:
                print(f"    * 错误: 处理图片时出现异常: {str(e)}")
                traceback.print_exc()
                retry_count += 1
        
        return success
    
    def translate_text(self, text, to_lang="zh"):
        """翻译文本"""
        if not has_translate_modules:
            # 如果没有翻译模块，直接返回空字符串
            print("警告: 翻译模块未找到，无法进行翻译")
            return ""
        else:
            # 使用翻译模块
            try:
                # 先尝试百度翻译
                if translate_baidu_spec:
                    try:
                        # 创建百度翻译器实例（从配置文件加载API密钥）
                        baidu_translator = translate_baidu.BaiduTranslator()
                        result = baidu_translator.translate(text, 'auto', to_lang)
                        if result:
                            return result
                    except Exception as e:
                        print(f"警告: 百度翻译失败，尝试腾讯翻译: {str(e)}")
                
                # 如果百度翻译失败，尝试腾讯翻译
                if translate_tencent_spec:
                    try:
                        # 创建腾讯翻译器实例（从配置文件加载API密钥）
                        tencent_translator = translate_tencent.TencentTranslator(source='auto', target=to_lang)
                        response = tencent_translator.translate(text)
                        json_response = json.loads(response)
                        if 'Response' in json_response and 'TargetText' in json_response['Response']:
                            return json_response['Response']['TargetText']
                    except Exception as e:
                        print(f"警告: 腾讯翻译失败: {str(e)}")
                
                # 如果两个都失败，返回空字符串
                print("警告: 所有翻译方式均失败，无法进行翻译")
                return ""
            except Exception as e:
                print(f"错误: 翻译时出现异常: {str(e)}")
                return ""
    
    def resize_image_for_excel(self, image_path, max_size=256):
        """调整图片大小用于Excel"""
        try:
            img = Image.open(image_path)
            # 计算原始图片的宽高比
            width, height = img.size
            img_scale = width / height
            
            # 固定行高为100pt (约133px)
            height_pt = 100  # 目标高度为 100 pt
            height_px = 100 * (4 / 3)  # 转换为像素单位
            
            # 固定高度，根据图片比例计算宽度
            new_height = height_px
            new_width = int(new_height * img_scale)
            
            # 调整大小
            img = img.resize((new_width, new_height), Image.LANCZOS)
            
            # 保存到内存
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            return buffer.getvalue(), new_width, new_height, img_scale, height_pt
        except Exception as e:
            print(f"错误: 调整图片大小时出现异常: {str(e)}")
            return None
    
    def batch_update_excel(self, successful_images):
        """批量更新Excel文件中的提示词和图片信息"""
        if not successful_images:
            print("没有需要更新的图片信息")
            return
            
        try:
            # 打开Excel文件
            print(f"打开Excel文件: {self.excel_path}")
            wb = openpyxl.load_workbook(self.excel_path)
            
            # 检查是否有"提示词"工作表
            if "提示词" not in wb.sheetnames:
                print(f"错误: Excel文件中没有'提示词'工作表")
                return
            
            ws = wb["提示词"]
            
            # 找到第一个空行
            start_row = 1
            while ws.cell(row=start_row, column=1).value is not None:
                start_row += 1
            print(f"在Excel中找到起始空行: 第{start_row}行")
            
            # 批量添加数据
            print(f"开始批量添加{len(successful_images)}个图片的提示词和信息")
            
            for i, image_info in enumerate(successful_images):
                current_row = start_row + i
                image_path = image_info["image_path"]
                prompt = image_info["prompt"]
                
                print(f"  - 处理第{i+1}/{len(successful_images)}个图片: {os.path.basename(image_path)}")
                
                # 添加英文提示词
                ws.cell(row=current_row, column=1, value=prompt)  # 英文提示词
                
                # 翻译成中文
                chinese_prompt = self.translate_text(prompt)
                if chinese_prompt:
                    print(f"    * 翻译成功")
                else:
                    print(f"    * 翻译失败，中文提示词为空")
                ws.cell(row=current_row, column=2, value=chinese_prompt)  # 中文提示词
                
                # 添加图片预览
                print(f"    * 调整图片大小并添加预览到Excel")
                img_data = self.resize_image_for_excel(image_path)
                if img_data:
                    img_data, new_width, new_height, img_scale, height_pt = img_data
                    img = openpyxl.drawing.image.Image(BytesIO(img_data))
                    cell = ws.cell(row=current_row, column=3)
                    
                    # 设置图片的新宽度和高度
                    img.width = new_width
                    img.height = new_height
                    
                    # 计算对应的列宽（字符单位）
                    height_px = height_pt * (4 / 3)
                    width_ch = height_px / 8
                    
                    # 设置单元格宽高
                    col_letter = openpyxl.utils.get_column_letter(3)
                    ws.column_dimensions[col_letter].width = width_ch
                    ws.row_dimensions[current_row].height = height_pt  # 行高单位为磅
                    
                    img.anchor = cell.coordinate
                    ws.add_image(img)
                    print(f"    * 图片预览添加成功")
                else:
                    print(f"    * 图片预览添加失败")
                
                # 添加图片路径
                ws.cell(row=current_row, column=4, value=image_path)  # 图片路径
            
            # 保存Excel文件
            print(f"保存Excel文件")
            wb.save(self.excel_path)
            print(f"Excel批量更新成功，共更新了{len(successful_images)}个图片的信息")
            
        except Exception as e:
            print(f"错误: 批量更新Excel文件时出现异常: {str(e)}")
            traceback.print_exc()

    def update_excel(self, image_path, success):
        """更新Excel文件中的提示词和图片信息（单个图片版本，已弃用）"""
        try:
            # 打开Excel文件
            print(f"    * 打开Excel文件: {self.excel_path}")
            wb = openpyxl.load_workbook(self.excel_path)
            
            # 检查是否有"提示词"工作表
            if "提示词" not in wb.sheetnames:
                print("    * 错误: Excel文件中没有'提示词'工作表")
                return
            
            ws = wb["提示词"]
            
            # 找到第一个空行
            row = 1
            while ws.cell(row=row, column=1).value is not None:
                row += 1
            print(f"    * 在Excel中找到空行: 第{row}行")
            
            # 添加数据
            print(f"    * 添加英文提示词到Excel")
            ws.cell(row=row, column=1, value=self.picture_prompt)  # 英文提示词
            
            # 翻译成中文
            print(f"    * 翻译提示词为中文")
            chinese_prompt = self.translate_text(self.picture_prompt)
            if chinese_prompt:
                print(f"    * 翻译成功，添加中文提示词到Excel")
            else:
                print(f"    * 翻译失败，中文提示词为空")
            ws.cell(row=row, column=2, value=chinese_prompt)  # 中文提示词
            
            # 添加图片预览
            print(f"    * 调整图片大小并添加预览到Excel")
            img_data = self.resize_image_for_excel(image_path)
            if img_data:
                img_data, new_width, new_height, img_scale, height_pt = img_data
                img = openpyxl.drawing.image.Image(BytesIO(img_data))
                cell = ws.cell(row=row, column=3)
                
                # 计算对应的列宽（字符单位）
                width_ch = new_width / 8  # 列宽单位为字符，像素单位需除以 8
                
                # 设置图片的新宽度和高度
                img.width = new_width
                img.height = new_height
                
                # 设置单元格宽高
                col_letter = openpyxl.utils.get_column_letter(3)
                ws.column_dimensions[col_letter].width = width_ch
                ws.row_dimensions[row].height = height_pt  # 行高单位为磅
                
                img.anchor = cell.coordinate
                ws.add_image(img)
                print(f"    * 图片预览添加成功")
            else:
                print(f"    * 图片预览添加失败")
            
            # 添加图片路径
            print(f"    * 添加图片路径到Excel: {image_path}")
            ws.cell(row=row, column=4, value=image_path)  # 图片路径
            
            # 保存Excel文件
            print(f"    * 保存Excel文件")
            wb.save(self.excel_path)
            print(f"    * Excel更新成功")
            
        except Exception as e:
            print(f"    * 错误: 更新Excel文件时出现异常: {str(e)}")
            traceback.print_exc()
    
    def update_step_status(self):
        """更新步骤工作表的完成结果"""
        try:
            # 打开Excel文件
            print(f"\n更新步骤工作表的完成结果...")
            print(f"打开Excel文件更新步骤状态: {self.excel_path}")
            wb = openpyxl.load_workbook(self.excel_path)
            
            # 检查是否有"步骤"工作表
            if "步骤" not in wb.sheetnames:
                print("错误: Excel文件中没有'步骤'工作表")
                return
            
            steps_ws = wb["步骤"]
            print("找到'步骤'工作表")
            
            # 查找图片描述生成步骤
            step_found = False
            for r in range(1, 20):  # 假设步骤不超过20行
                step_name = steps_ws.cell(row=r, column=1).value
                if step_name == "图片描述生成":
                    step_found = True
                    print(f"找到'图片描述生成'步骤，位于第{r}行")
                    
                    # 更新时间
                    current_time = time.strftime("%Y-%m-%d %H:%M:%S")
                    steps_ws.cell(row=r, column=3, value=current_time)
                    print(f"更新完成时间: {current_time}")
                    
                    # 更新状态 - 改进的判断逻辑
                    status = ""
                    
                    # 检查是否有前置工作失败的情况
                    if not hasattr(self, 'client') or not hasattr(self, 'workflow') or not self.excel_path or not self.gemini_folder_path:
                        status = "失败"
                        print("更新状态: 失败 (前置工作未完成)")
                    # 检查是否所有图片都成功处理
                    elif not self.failed_images:  
                        status = "成功"
                        print("更新状态: 成功 (所有图片处理成功)")
                    # 处理部分图片失败的情况
                    else:  
                        # 格式化失败信息：文件名1：失败，文件名2：失败，...
                        failed_info = "，".join([f"{name}：失败" for name in self.failed_images])
                        status = failed_info
                        print(f"更新状态: 部分失败 ({len(self.failed_images)}个文件失败)")
                    
                    # 更新状态到Excel
                    steps_ws.cell(row=r, column=4, value=status)
                    break
            
            if not step_found:
                print("警告: 未在'步骤'工作表中找到'图片描述生成'步骤")
            
            # 保存Excel文件
            print("保存Excel文件")
            wb.save(self.excel_path)
            print("步骤状态更新完成")
            
        except Exception as e:
            print(f"错误: 更新步骤状态时出现异常: {str(e)}")
            traceback.print_exc()


def main():
    print("\n========================================")
    print("   Lora训练 - 图片描述生成工具 (Gemini)   ")
    print("========================================\n")
    
    # 检查命令行参数
    if len(sys.argv) < 2:
        print("用法: python #Lora_2_画面描述-Gemini.py <resize文件夹路径>")
        print("示例: python #Lora_2_画面描述-Gemini.py D:\\训练项目\\resize")
        return
    
    resize_folder_path = sys.argv[1]
    print(f"输入参数: resize文件夹路径 = {resize_folder_path}")
    
    if not os.path.exists(resize_folder_path):
        print(f"错误: 路径不存在: {resize_folder_path}")
        return
    
    # 显示开始时间
    start_time = time.time()
    print(f"开始时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    try:
        # 创建图片描述生成器并处理图片
        print("初始化图片描述生成器...")
        generator = ImageDescriptionGenerator(resize_folder_path)
        generator.process_images()
        
        # 计算总耗时
        elapsed_time = time.time() - start_time
        minutes, seconds = divmod(elapsed_time, 60)
        print(f"\n处理完成! 总耗时: {int(minutes)}分{int(seconds)}秒")
        print(f"结束时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        print("========================================")
    except Exception as e:
        print(f"\n错误: 执行过程中发生异常: {str(e)}")
        traceback.print_exc()
        print("\n程序异常终止!")
        print("========================================")


if __name__ == "__main__":
    main()