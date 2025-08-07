import os
import sys
import time
import socket
import shutil
import openpyxl
import subprocess
import importlib.util
import pynvml
import pygetwindow as gw
import pyautogui
import pyperclip
import psutil
import datetime
import traceback

# 导入自定义工具包
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.ChromeManager import ChromeManager

# 定义训练集类型和地址
CHARACTER_DIR = "E:\\Design\\Character"
STYLES_DIR = "E:\\Design\\Styles"


class LoraTrainer:
    def __init__(self, project_path):
        self.project_path = project_path
        self.project_type = None
        self.project_name = None
        self.excel_path = None
        self.toml_template = None
        self.toml_path = None
        self.chrome_manager = None
        self.driver = None
        self.wait = None
        self.training_start_time = None
        self.training_end_time = None
        self.training_result = "失败"  # 默认为失败，成功时会更新
        
        # 初始化NVML
        pynvml.nvmlInit()
    
    def __del__(self):
        # 关闭NVML
        try:
            pynvml.nvmlShutdown()
        except:
            pass
    
    def detect_project_info(self):
        """检测项目类型和名称"""
        if self.project_path.startswith(CHARACTER_DIR):
            self.project_type = "Character"
            self.project_name = os.path.basename(self.project_path)
        elif self.project_path.startswith(STYLES_DIR):
            self.project_type = "Styles"
            self.project_name = os.path.basename(self.project_path)
        else:
            raise ValueError(f"项目路径 {self.project_path} 不在有效的训练集目录下")
        
        print(f"项目类型: {self.project_type}, 项目名称: {self.project_name}")
        
        # 查找训练信息Excel文件
        self.excel_path = os.path.join(self.project_path, "训练信息.xlsx")
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"未找到训练信息Excel文件: {self.excel_path}")
    
    def read_training_template(self):
        """从Excel文件中读取训练模板"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            steps_sheet = wb["步骤"]
            
            # 查找训练模板行
            template_name = None
            for row in range(2, steps_sheet.max_row + 1):
                step_name = steps_sheet.cell(row=row, column=1).value
                if step_name == "训练模板":
                    template_name = steps_sheet.cell(row=row, column=2).value
                    break
            
            if not template_name:
                raise ValueError("未在Excel文件中找到训练模板名称")
            
            print(f"训练模板: {template_name}")
            
            # 查找对应的toml文件
            toml_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "toml")
            self.toml_template = os.path.join(toml_dir, f"{template_name}.toml")
            
            if not os.path.exists(self.toml_template):
                # 修复：确保在发生错误时设置训练结果为失败
                self.training_result = "失败"
                raise FileNotFoundError(f"未找到训练模板文件: {self.toml_template}")
            
            return template_name
        except Exception as e:
            print(f"读取训练模板时出错: {str(e)}")
            traceback.print_exc()
            # 修复：确保在发生异常时设置训练结果为失败
            self.training_result = "失败" 
            return None
    
    def modify_toml_file(self):
        """修改toml配置文件"""
        try:
            # 读取原始toml文件内容
            with open(self.toml_template, "r", encoding="utf-8") as f:
                toml_content = f.read()
            
            # 直接使用原始模板文件路径
            self.toml_path = self.toml_template
            
            # 修改参数
            # 修改train_data_dir
            train_data_dir = f"E:/Design/{self.project_type}/{self.project_name}/gemini"
            toml_content = self.replace_toml_param(toml_content, "train_data_dir", train_data_dir)
            
            # 修改output_name
            toml_content = self.replace_toml_param(toml_content, "output_name", self.project_name)
            
            # 修改output_dir
            output_dir = f"E:/Design/loras/{self.project_name}"
            
            # 检查输出目录是否存在，如果不存在则创建
            output_dir_windows_path = output_dir.replace("/", "\\")
            if not os.path.exists(output_dir_windows_path):
                try:
                    os.makedirs(output_dir_windows_path, exist_ok=True)
                    print(f"已创建输出目录: {output_dir_windows_path}")
                except Exception as e:
                    print(f"创建输出目录时出错: {str(e)}")
                    traceback.print_exc()
            else:
                print(f"输出目录已存在: {output_dir_windows_path}")
                
            toml_content = self.replace_toml_param(toml_content, "output_dir", output_dir)
            
            # 修改log_prefix
            toml_content = self.replace_toml_param(toml_content, "log_prefix", self.project_name)
            
            # 修改log_tracker_name
            toml_content = self.replace_toml_param(toml_content, "log_tracker_name", self.project_name)
            
            # 直接保存到原始toml文件
            with open(self.toml_path, "w", encoding="utf-8") as f:
                f.write(toml_content)
            
            print(f"已修改并保存toml配置文件: {self.toml_path}")
            return True
        except Exception as e:
            print(f"修改toml文件时出错: {str(e)}")
            traceback.print_exc()
            return False
    
    def replace_toml_param(self, content, param_name, new_value):
        """替换toml文件中的参数值"""
        import re
        pattern = fr"({param_name}\s*=\s*)([\"'].*?[\"']|[^\"'\n]+)"
        # 处理new_value中的反斜杠，确保它们被正确转义
        escaped_value = new_value.replace('\\', '\\\\')
        replacement = f"\\g<1>\"{escaped_value}\""
        return re.sub(pattern, replacement, content)
    
    def check_training_server(self):
        """检查训练服务器是否运行"""
        try:
            # 检查端口是否开放
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(2)
            result = sock.connect_ex(("127.0.0.1", 28000))
            sock.close()
            
            if result == 0:
                print("训练服务器已运行")
                return True
            else:
                print("训练服务器未运行，尝试启动...")
                # 使用os.startfile直接打开批处理文件
                os.startfile("F:\\LoraTrain\\A启动脚本.bat")
                
                # 等待服务器启动
                start_time = time.time()
                while time.time() - start_time < 120:  # 最多等待2分钟
                    time.sleep(10)
                    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock.settimeout(2)
                    result = sock.connect_ex(("127.0.0.1", 28000))
                    sock.close()
                    if result == 0:
                        print("训练服务器已成功启动")
                        return True
                
                print("训练服务器启动超时")
                return False
        except Exception as e:
            print(f"检查训练服务器时出错: {str(e)}")
            traceback.print_exc()
            return False
    
    def kill_chrome_processes(self):
        """结束所有Chrome进程"""
        try:
            for proc in psutil.process_iter(['pid', 'name']):
                if proc.info['name'] and 'chrome.exe' in proc.info['name'].lower():
                    try:
                        proc.kill()
                    except:
                        pass
            time.sleep(2)  # 等待进程结束
            print("已结束所有Chrome进程")
        except Exception as e:
            print(f"结束Chrome进程时出错: {str(e)}")
    
    def init_chrome_manager(self):
        """初始化Chrome管理器"""
        try:
            # 结束所有Chrome进程
            self.kill_chrome_processes()
            
            # 配置Chrome管理器
            config = {
                'chrome_path': r'D:\Code\.module\chrome\chrome-win-115\chrome.exe',
                'chromedriver_path': r'D:\Code\.module\chrome-driver\115\chromedriver.exe',
                'user_data_dir': r"D:\Code\.module\chrome-cache\chrome-cache-115\AutomationProfile",
                'headless': False,
                'enable_extensions': False,
                'using_proxy': False,
                'enable_images': True
            }
            
            self.chrome_manager = ChromeManager(config)
            self.driver, self.wait = self.chrome_manager.open_url("http://127.0.0.1:28000/lora/flux.html")
            time.sleep(5)
            if not self.driver:
                raise Exception("无法打开训练页面")
            
            print("已成功打开训练页面")
            return True
        except Exception as e:
            print(f"初始化Chrome管理器时出错: {str(e)}")
            traceback.print_exc()
            return False
    
    def start_training(self):
        """开始训练"""
        try:
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support import expected_conditions as EC
            import time
            
            
            # 点击导入配置文件按钮 - 使用CSS选择器
            import_button = self.driver.find_element(By.CSS_SELECTOR, "#app > div > div.example-container > div.right-container > div:nth-child(4) > div:nth-child(2) > button")
            import_button.click()
            time.sleep(5)
            
            # 将toml文件路径复制到剪贴板
            pyperclip.copy(self.toml_path)
            
            # 模拟键盘操作粘贴路径并回车
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(1)
            pyautogui.press('enter')
            time.sleep(3)
            
            # 点击开始训练按钮 - 使用CSS选择器
            start_button = self.driver.find_element(By.CSS_SELECTOR, "#app > div > div.example-container > div.right-container > div:nth-child(6) > div:nth-child(1) > button")
            start_button.click()
            
            print("已点击开始训练按钮")
            self.training_start_time = datetime.datetime.now()
            
            return True
        except Exception as e:
            print(f"开始训练时出错: {str(e)}")
            traceback.print_exc()
            return False
    
    def get_gpu_utilization(self):
        """获取GPU使用率"""
        try:
            handle = pynvml.nvmlDeviceGetHandleByIndex(0)
            utilization = pynvml.nvmlDeviceGetUtilizationRates(handle).gpu
            return utilization
        except pynvml.NVMLError as e:
            print(f"获取GPU使用率时出错: {e}")
            return None
    
    def get_gpu_memory_usage(self):
        """获取GPU内存使用量(GB)"""
        try:
            handle = pynvml.nvmlDeviceGetHandleByIndex(0)
            memory_info = pynvml.nvmlDeviceGetMemoryInfo(handle)
            used_memory = memory_info.used / (1024 ** 3)
            return used_memory
        except pynvml.NVMLError as e:
            print(f"获取GPU内存使用量时出错: {e}")
            return None
    
    def find_bat_window(self):
        """尝试查找批处理窗口，支持多种窗口标题匹配方式
        返回找到的窗口对象，如果未找到则返回None
        """
        # 可能的窗口标题列表
        possible_titles = [
            "A启动脚本.bat",  # 直接使用批处理文件名作为标题
            "C:\\WINDOWS\\system32\\cmd.exe"  # 使用cmd.exe作为标题（批处理脚本运行时的实际窗口标题）
        ]
        
        # 尝试每一种可能的标题
        for title in possible_titles:
            windows = gw.getWindowsWithTitle(title)
            if windows:
                print(f"找到窗口: {title}，共 {len(windows)} 个")
                return windows[0]  # 返回第一个匹配的窗口
        
        # 如果所有标题都未找到匹配的窗口
        return None
        
    def activate_bat_window_and_press_enter(self):
        """激活批处理窗口并按回车，支持多种窗口标题匹配方式"""
        try:
            # 尝试查找批处理窗口
            window = self.find_bat_window()
            if window:
                window.activate()
                pyautogui.press('enter')
                print(f"已激活批处理窗口 '{window.title}' 并按回车")
            else:
                print("未找到批处理窗口，请先手动启动批处理文件")
        except Exception as e:
            print(f"激活窗口时出错: {e}")
    
    def monitor_training(self):
        """监控训练过程"""
        try:
            # 首先检查是否开始运行
            print("开始监控训练过程...")
            start_time = time.time()
            training_started = False
            
            # 检查是否开始运行
            while time.time() - start_time < 180:  # 最多等待3分钟
                utilization = self.get_gpu_utilization()
                memory_usage = self.get_gpu_memory_usage()
                
                print(f"GPU使用率: {utilization}%, 内存使用量: {memory_usage:.2f}GB")
                
                if utilization > 50 or (memory_usage and memory_usage > 15):
                    training_started = True
                    print("训练已开始运行")
                    break
                
                time.sleep(10)
            
            if not training_started:
                print("训练未开始运行，尝试重启...")
                self.activate_bat_window_and_press_enter()
                
                # 再次等待训练开始
                start_time = time.time()
                while time.time() - start_time < 180:  # 最多等待3分钟
                    utilization = self.get_gpu_utilization()
                    memory_usage = self.get_gpu_memory_usage()
                    
                    print(f"GPU使用率: {utilization}%, 内存使用量: {memory_usage:.2f}GB")
                    
                    if utilization > 50 or (memory_usage and memory_usage > 15):
                        training_started = True
                        print("训练已开始运行")
                        break
                    
                    time.sleep(10)
            
            if not training_started:
                print("训练无法启动，退出监控")
                return False
            
            # 持续监控训练过程
            inactive_count = 0
            while True:
                utilization_records = []
                memory_records = []
                
                # 收集30秒内的数据
                for _ in range(6):  # 每5秒一次，共30秒
                    utilization = self.get_gpu_utilization()
                    memory_usage = self.get_gpu_memory_usage()
                    
                    if utilization is not None:
                        utilization_records.append(utilization)
                    if memory_usage is not None:
                        memory_records.append(memory_usage)
                    
                    time.sleep(5)
                
                # 计算平均值
                avg_utilization = sum(utilization_records) / len(utilization_records) if utilization_records else 0
                avg_memory = sum(memory_records) / len(memory_records) if memory_records else 0
                
                print(f"平均GPU使用率: {avg_utilization:.2f}%, 平均内存使用量: {avg_memory:.2f}GB")
                
                # 检查是否需要重启
                # GPU使用率<50%且内存>15GB：认为训练卡住，需要重启
                if all(util < 30 for util in utilization_records) and any(mem > 15 for mem in memory_records):
                    inactive_count += 1
                    print(f"检测到训练可能卡住，计数: {inactive_count}/6")
                    if inactive_count >= 6:  # 连续6次检测到异常（约3分钟）才重启
                        print("确认训练卡住，尝试重启...")
                        self.activate_bat_window_and_press_enter()
                        inactive_count = 0  # 重置计数器
                # GPU使用率<10%且内存<3GB：可能已停止，连续3分钟后确认
                elif all(util < 20 for util in utilization_records) and all(mem < 3 for mem in memory_records):
                    inactive_count += 1
                    print(f"检测到训练可能已停止，计数: {inactive_count}/10")
                    if inactive_count >= 10:  # 连续10次检测到低活动（约5分钟）
                        print("训练已停止，退出监控")
                        return False
                else:
                    inactive_count = 0  # 重置计数器
        
        except Exception as e:
            print(f"监控训练过程时出错: {str(e)}")
            traceback.print_exc()
            return False
            
    def check_training_completion(self):
        """检查训练是否真正完成"""
        try:
            # 检查是否存在训练完成的模型文件
            safetensors_path = f"E:/Design/loras/{self.project_name}/{self.project_name}.safetensors"
            if os.path.exists(safetensors_path):
                print(f"发现训练完成的模型文件: {safetensors_path}")
                self.training_end_time = datetime.datetime.now()
                self.training_result = "成功"
                return True
            else:
                print(f"未找到训练完成的模型文件: {safetensors_path}")
                self.training_result = "失败"
                return False
        except Exception as e:
            print(f"检查训练完成状态时出错: {str(e)}")
            traceback.print_exc()
            self.training_result = "失败"
            return False
    
    def update_excel(self):
        """更新Excel文件中的训练结果"""
        try:
            # 打开Excel文件
            wb = openpyxl.load_workbook(self.excel_path)
            steps_sheet = wb["步骤"]
            
            # 查找模型Lora训练步骤
            for row in range(2, steps_sheet.max_row + 1):
                step_name = steps_sheet.cell(row=row, column=1).value
                if step_name == "模型Lora训练":
                    # 更新完成时间
                    if self.training_end_time:
                        steps_sheet.cell(row=row, column=3, value=self.training_end_time.strftime("%Y-%m-%d %H:%M:%S"))
                    else:
                        steps_sheet.cell(row=row, column=3, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    
                    # 更新完成结果
                    steps_sheet.cell(row=row, column=4, value=self.training_result)
                    break
            
            # 保存Excel文件
            wb.save(self.excel_path)
            print(f"已更新Excel文件: {self.excel_path}")
            return True
        except Exception as e:
            print(f"更新Excel文件时出错: {str(e)}")
            traceback.print_exc()
            return False
    
    def run(self):
        """运行训练流程"""
        try:
            # 检测项目信息
            self.detect_project_info()
            
            # 读取训练模板
            template_name = self.read_training_template()
            if not template_name:
                print("无法读取训练模板，退出")
                # 修复：确保在退出前设置训练结果为失败
                self.training_result = "失败"
                self.update_excel()  # 更新失败结果
                return False
            
            # 修改toml文件
            if not self.modify_toml_file():
                print("无法修改toml文件，退出")
                # 修复：确保在退出前设置训练结果为失败
                self.training_result = "失败"
                self.update_excel()  # 更新失败结果
                return False
            
            # 检查训练服务器
            if not self.check_training_server():
                print("无法启动训练服务器，退出")
                self.training_result = "失败"  # 确保设置为失败
                self.update_excel()  # 更新失败结果
                return False
            
            # 初始化Chrome管理器
            if not self.init_chrome_manager():
                print("无法初始化Chrome管理器，退出")
                self.training_result = "失败"  # 确保设置为失败
                self.update_excel()  # 更新失败结果
                return False
            
            # 开始训练
            if not self.start_training():
                print("无法开始训练，退出")
                self.update_excel()  # 更新失败结果
                return False
            
            # 监控训练过程
            monitor_result = self.monitor_training()
            
            # 无论监控结果如何，都检查训练是否真正完成
            training_success = self.check_training_completion()
            
            # 更新Excel文件
            self.update_excel()
            
            return training_success
        except Exception as e:
            print(f"运行训练流程时出错: {str(e)}")
            traceback.print_exc()
            # 确保在发生异常时检查训练是否完成
            self.check_training_completion()
            self.update_excel()  # 更新结果
            return False


def main():
    # 检查命令行参数
    if len(sys.argv) < 2:
        print("用法: python #Lora_4_模型训练.py <项目路径>")
        return 1
    
    project_path = sys.argv[1]
    if not os.path.exists(project_path):
        print(f"错误: 项目路径不存在: {project_path}")
        return 1
    
    # 创建训练器并运行
    trainer = LoraTrainer(project_path)
    success = trainer.run()
    
    if success:
        print("训练成功完成")
        return 0
    else:
        print("训练失败")
        return 1


if __name__ == "__main__":
    sys.exit(main())