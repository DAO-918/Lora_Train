"""
以下是部分json文件的内容，记录了所有模型的信息：
{
  "Lora\\SD_XL_角色\\25D_XL_角色_marie rose.safetensors": {
    "name": "25D_XL_角色_marie rose",
    "type": "",
    "url": "https://civitai.com/models/128594?modelVersionId=385001",
    "description": "=== 从Civitai抓取的描述 ===\nI've trained two sets of clothing. The trigger words are as follows\n\ntype A: pantyhose, sweater, jacket\n\ntype B: shirt, skirt, fishnet thighhighs\n\n\nV3.0\n\nThe previous 1.0 version of Marie was really too ugly, so I collected a large amount of new material and made the latest XL version, which took a lot of time... I hope everyone can provide more image feedback to build the SDXL community ecosystem and let more people know the advantages of XL...\n\n\n现在网上的ai玛丽实在是太丑了不忍直视，因此希望大家用这个模型创造出更多可爱的还原的玛丽，更多的优质返图就是对我最大的支持，也希望和更多的朋友讨论如何生成更好看的玛丽。可以关注我的p站和X账号，以后会有更多的图片创作以及优质模型放出...",
    "trigger_words": "marie rose",
    "hash": "bdfcc89aa58d85c7909b4b85f813b50797cc3c0a9120ab8b17a40c18d4bfee01",
    "is_favorite": false,
    "last_modified": 1734931371.922711
  },
python，全程使用openpyxl
第一层键名加上folder_path就是文件的存储路径,该路径下有同名的npg文件和关于该模型的json信息，示例如下：
{
    "description": "description",
    "sd version": "Flux",
    "activation text": "marie rose",
    "preferred weight": 0.76,
    "negative text": "Negative prompt\n",
    "notes": "Notes"
}
第二层键名包括了：name,type,url,description,trigger_words,hash,is_favorite,last_modified

现在要求将json的信息转换成execl的格式
"Lora\\SD_XL_角色\\25D_XL_角色_marie rose.safetensors"，"Lora"是写入sheet的名称，"SD_XL_角色"是模型的文件夹名，"25D_XL_角色_marie rose"是模型的文件名，safetensors是文件的后缀名
同时要获取该模型路径下的同名的npg文件路径和同名的json文件(关于该模型的json信息)，要判断是否存在png,没有则不写入img_path和插入图片；要判断是否存在png，没有则preferred weight,negative text,notes不写入
表格从A列开始依次是：name,文件夹名,type,url,img_path,插入图片,description,notes,trigger_words,preferred weight,negative text,hash,is_favorite,last_modified
"""

"""
headers = [
                "文件名", "原名", "文件夹名", "类型", "风格", "用途", "版本", "url", "图片路径", "图片预览", 
                "描述", "触发词", "可选组合", "notes", "默认权重", "权重范围", "否定提示词", 
                "hash","喜爱", "修改时间"
            ]
使用python和openpyxl，我会修改表格中的内容，新的文件名按照 "类型_用途_风格_原名_版本" 的格式，如果有空值则不写入到新文件名中。
对比原文件名(即表格中的文件名)，如果不一致，则修改模型的文件名以及对应图片文件名和JSON文件名，如果一致，则不修改。
下面是我写的一部分代码，请另外写出一个检查并修改文件名的方法。
"""

"""
model_info.josn的部分内容：
{
  "Lora\\SD_XL_角色\\25D_XL_角色_marie rose.safetensors": {
    "name": "25D_XL_角色_marie rose",
    "type": "",
    "url": "https://civitai.com/models/128594?modelVersionId=385001",
    "description": "=== 从Civitai抓取的描述 ===\nI've trained two sets of clothing. The trigger words are as follows\n\ntype A: pantyhose, sweater, jacket\n\ntype B: shirt, skirt, fishnet thighhighs\n\n\nV3.0\n\nThe previous 1.0 version of Marie was really too ugly, so I collected a large amount of new material and made the latest XL version, which took a lot of time... I hope everyone can provide more image feedback to build the SDXL community ecosystem and let more people know the advantages of XL...\n\n\n现在网上的ai玛丽实在是太丑了不忍直视，因此希望大家用这个模型创造出更多可爱的还原的玛丽，更多的优质返图就是对我最大的支持，也希望和更多的朋友讨论如何生成更好看的玛丽。可以关注我的p站和X账号，以后会有更多的图片创作以及优质模型放出...",
    "trigger_words": "marie rose",
    "hash": "bdfcc89aa58d85c7909b4b85f813b50797cc3c0a9120ab8b17a40c18d4bfee01",
    "is_favorite": false,
    "last_modified": 1734931371.922711
  },

headers = [
                "文件名", "原名", "文件夹名", "类型", "风格", "用途", "版本", "url", "图片路径", "图片预览", 
                "描述", "触发词", "可选组合", "notes", "默认权重", "权重范围", "否定提示词", 
                "hash","喜爱", "修改时间"
            ]
将excel的数据重新写到model_info.json和模型对应同名的json文件中。
其中model_info.json的键名是工作表sheet名称+文件名+文件名，需要写入的表格中包含的所有数据，除了原来model_info.json有的值，其他的键名按表格中的列名命名


模型对应同名的json文件的示例如下：
{
    "description": "=== 从Civitai抓取的描述 ===\nI've trained two sets of clothing. The trigger words are as follows\n\ntype A: pantyhose, sweater, jacket\n\ntype B: shirt, skirt, fishnet thighhighs\n\n\nV3.0\n\nThe previous 1.0 version of Marie was really too ugly, so I collected a large amount of new material and made the latest XL version, which took a lot of time... I hope everyone can provide more image feedback to build the SDXL community ecosystem and let more people know the advantages of XL...\n\n\n现在网上的ai玛丽实在是太丑了不忍直视，因此希望大家用这个模型创造出更多可爱的还原的玛丽，更多的优质返图就是对我最大的支持，也希望和更多的朋友讨论如何生成更好看的玛丽。可以关注我的p站和X账号，以后会有更多的图片创作以及优质模型放出...",
    "sd version": "Flux",
    "activation text": "marie rose",
    "preferred weight": 0.76,
    "negative text": "Negative prompt\n",
    "notes": "Notes"
}
也一样写入的表格中包含的所有数据，除了原来有的值，其他的键名按表格中的列名命名

下面是我写的部分代码，仅供参考，另外写一个方法完成我上面提到的要求：
"""

"""
模仿以下代码风格，不需要考虑代码效率优化，保持简单易懂
模仿代码写法，另外写一个方法，当新文件夹列中有值时需要将模型、json文件和图片都移动到新文件夹下

"""

import logging
import sys
import io
import os
import time
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import openpyxl.utils
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl import Workbook


def json_to_execl(folder_path):
    json_name = "model_info.json"
    json_path = os.path.join(folder_path, json_name)
    with open(json_path, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    
    excel_name = "model_info.xlsx"
    excel_path = os.path.join(folder_path, excel_name)
    wb = load_workbook(excel_path)
    
    sheets = wb.sheetnames
    # 先遍历所有工作表清空已有图片
    for sheet in sheets:
        ws = wb[sheet]
        # 获取图片的列表
        images = list(ws._images)  # 创建一个副本以避免在迭代时修改列表长度
        for drawing in images:
            ws._images.remove(drawing)
    
    # 删除默认生成的工作表（通常名为'Sheet'）
    default_sheet_name = 'Sheet'
    for sheet in wb.sheetnames:
        if default_sheet_name in sheet:
            default_sheet = wb[sheet]
            wb.remove(default_sheet)
    
    for file_path, model_info in json_data.items():
        # 分解路径和文件信息
        split_list = file_path.split('\\')
        if len(split_list) == 2:
            sheet_name = split_list[0]
            folder_name = ''
            file_name_ext = split_list[1]
        elif len(split_list) == 3:
            sheet_name = split_list[0]
            folder_name = split_list[1]
            file_name_ext = split_list[2]
        file_name, file_ext = os.path.splitext(file_name_ext)
        
        # 获取相关路径  
        model_dir = os.path.join(folder_path,sheet_name)
        if bool(folder_name):
            model_dir = os.path.join(model_dir,folder_name)
        img_path = os.path.join(model_dir, file_name + '.png')
        model_json_path = os.path.join(model_dir, file_name + '.json')
        
        # 判断模型文件是否存在
        model_path = os.path.join(model_dir, file_name_ext)
        if not os.path.exists(model_path):
            continue
        
        # 读取模型 JSON 信息
        model_extra_info = {}
        if os.path.exists(model_json_path):
            with open(model_json_path, 'r', encoding='utf-8') as f:
                model_extra_info = json.load(f)
        
        # 检查文件是否存在
        img_exists = os.path.exists(img_path)
        json_exists = os.path.exists(model_json_path)
        
        # 如果工作表不存在，则创建
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            # 写入表头
            headers = [
                "文件名", "拓展名", "原名", "文件夹名", "新文件夹", "类型", "风格", "用途", "版本", "url", "图片路径", "图片预览", 
                "描述", "SD Link", "特指词", "主描述词", "触发词", "可选形象", "可选服装", "notes", "默认权重", "权重范围", "否定提示词", 
                "hash","喜爱", "修改时间"
            ]
            ws.append(headers)
        else:
            ws = wb[sheet_name]
        
        # 检查数据是否存在
        file_name_exsit = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == file_name:
                file_name_exsit = True
                break
        if file_name_exsit:
            continue
        
        # 写入行数据
        row_data = [
            file_name, #1 文件名
            file_ext,  #2 拓展名
            model_info.get("pname", file_name), #3 原名 如果不这样写第一次写入这一行为空，会影响后面程序的运行，在rename_filenames中旧文件名是有值的，而新文件名是空的，导致更新文件名直接变成空
            folder_name, #4 文件夹名
            "", #5 新文件夹
            model_info.get("type", ""), #6 类型 会写入模型json的sd version
            model_info.get("风格", ""), #7 风格
            model_info.get("用途", ""), #8 用途
            model_info.get("版本", ""), #9 版本
            model_info.get("url", ""), #10 url
            img_path, #11 图片路径 不判断图片是否存在  if img_exists else "" 填充时再判断
            "", #12 图片预览
            str(model_info.get("description", "")).replace('=', ''), #13 描述 会写入模型json的description
            model_info.get("SD Link", ""), # 14 SD Link
            model_info.get("specific_words", ""), # 15 特指词
            model_info.get("main_words", ""), #16 主描述词
            model_info.get("trigger_words", ""), #17 触发词 会写入模型json的activation text
            model_info.get("可选形象", ""), #18 可选形象
            model_info.get("可选服装", ""), #18 可选服装
            model_extra_info.get("notes", "") if json_exists else "", #19 notes 会写入模型json的notes
            model_extra_info.get("preferred weight", "") if json_exists else "", #20 默认权重 会写入模型json的preferred weight
            model_info.get("权重范围", ""), #21 权重范围
            model_extra_info.get("negative text", "") if json_exists else "", #22 否定提示词 会写入模型json的negative text
            model_info.get("hash", ""), #23 hash
            model_info.get("is_favorite", ""), #24 喜爱
            model_info.get("last_modified", "") #25 修改时间
        ]
        ws.append(row_data)
    
    wb.save(excel_path)


def rename_filenames(folder_path):
    excel_name = "model_info.xlsx"
    excel_path = os.path.join(folder_path, excel_name)
    wb = load_workbook(excel_path)
    
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        
        # 获取表头索引
        headers = [
                "文件名", "拓展名", "原名", "文件夹名", "新文件夹", "类型", "风格", "用途", "版本", "url", "图片路径", "图片预览", 
                "描述", "SD Link", "特指词", "主描述词", "触发词", "可选形象", "可选服装", "notes", "默认权重", "权重范围", "否定提示词", 
                "hash","喜爱", "修改时间"
            ]
        header_index = {header: idx for idx, header in enumerate(headers)}
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total_rows = len(rows)  
        columns = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        index_文件名 = columns.index('文件名')+1
        index_拓展名 = columns.index('拓展名')+1
        index_原名 = columns.index('原名')+1
        index_文件夹名 = columns.index('文件夹名')+1
        index_图片路径 = columns.index('图片路径')+1
        index_类型 = columns.index('类型')+1
        index_风格 = columns.index('风格')+1
        index_用途 = columns.index('用途')+1
        index_版本 = columns.index('版本')+1
        
        for row in ws.iter_rows(min_row=2):
            # 类型_用途_风格_原名_版本
            类型 = row[index_类型-1].value
            用途 = row[index_用途-1].value
            风格 = row[index_风格-1].value
            原名 = row[index_原名-1].value
            版本 = row[index_版本-1].value
            new_filename_parts = [类型, 用途, 风格, 原名, 版本]
            # bool 是一个内置函数，它会将每个元素转换成布尔值（True 或 False）。
            # 如果元素是 None、空字符串 ""、数字 0、空列表等假值，它会返回 False，否则返回 True
            # 如果 new_filename_parts 是 ["类型", "", "风格", None, "版本"]，那么 filter(bool, new_filename_parts) 会返回 ["类型", "风格", "版本"]
            new_filename = "_".join(filter(bool, new_filename_parts))
            old_filename = row[index_文件名-1].value
            
            文件夹名 = row[index_文件夹名-1].value
            if 文件夹名:
                file_folder = os.path.join(folder_path, sheet, 文件夹名)
            else:
                file_folder = os.path.join(folder_path, sheet)
                
            if not bool(old_filename) or not bool(new_filename):
                #print("至少有一个路径是假值")
                continue
                
            if old_filename != new_filename:
                try:
                    拓展名 = row[index_拓展名-1].value
                    old_model_path = os.path.join(file_folder, f'{old_filename}{拓展名}')
                    new_model_path = os.path.join(file_folder, f'{new_filename}{拓展名}')
                    #print(os.path.exists(old_model_path))
                    if os.path.exists(old_model_path):
                        os.rename(old_model_path, new_model_path)
                except Exception as e:
                    print(e)
                
                try:
                    图片路径 = row[index_图片路径-1].value
                    old_image_path = os.path.join(file_folder, f'{old_filename}.png')
                    new_image_path = os.path.join(file_folder, f'{new_filename}.png')
                    # 图片要存在才能修改
                    if 图片路径:
                        if os.path.exists(old_image_path):
                            os.rename(old_image_path, new_image_path)
                            row[index_图片路径-1].value = new_image_path
                except Exception as e:
                    print(e)
                
                try:
                    old_json_path = os.path.join(file_folder, f'{old_filename}.json')
                    new_json_path = os.path.join(file_folder, f'{new_filename}.json')
                    if os.path.exists(old_json_path):
                        os.rename(old_json_path, new_json_path)
                except Exception as e:
                    print(e)
                
                # 可行的，本质上是cell数据格式
                row[index_文件名-1].value = new_filename
    
    # 保存修改后的表格
    wb.save(excel_path)

def move_to_newfolder(folder_path):
    excel_name = "model_info.xlsx"
    excel_path = os.path.join(folder_path, excel_name)
    wb = load_workbook(excel_path)
    
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        
        # 获取表头索引
        columns = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        index_文件名 = columns.index('文件名') + 1
        index_拓展名 = columns.index('拓展名') + 1
        index_文件夹名 = columns.index('文件夹名') + 1
        index_新文件夹 = columns.index('新文件夹') + 1
        index_图片路径 = columns.index('图片路径') + 1
        
        for row in ws.iter_rows(min_row=2):
            文件名 = row[index_文件名 - 1].value
            拓展名 = row[index_拓展名 - 1].value
            文件夹名 = row[index_文件夹名 - 1].value
            新文件夹 = row[index_新文件夹 - 1].value
            图片路径 = row[index_图片路径 - 1].value
            
            # 原始文件夹路径
            if 文件夹名:
                old_folder = os.path.join(folder_path, sheet, 文件夹名)
            else:
                old_folder = os.path.join(folder_path, sheet)
            
            # 新文件夹路径
            if 新文件夹:
                new_folder = os.path.join(folder_path, sheet, 新文件夹)
                if not os.path.exists(new_folder):
                    os.makedirs(new_folder)
                
                # 移动模型文件
                try:
                    old_model_path = os.path.join(old_folder, f"{文件名}{拓展名}")
                    new_model_path = os.path.join(new_folder, f"{文件名}{拓展名}")
                    if os.path.exists(old_model_path):
                        os.rename(old_model_path, new_model_path)
                except Exception as e:
                    print(e)
                
                # 移动图片文件
                try:
                    old_image_path = os.path.join(old_folder, f"{文件名}.png")
                    new_image_path = os.path.join(new_folder, f"{文件名}.png")
                    if 图片路径:
                        if os.path.exists(old_image_path):
                            os.rename(old_image_path, new_image_path)
                except Exception as e:
                    print(e)
                
                # 移动 JSON 文件
                try:
                    old_json_path = os.path.join(old_folder, f"{文件名}.json")
                    new_json_path = os.path.join(new_folder, f"{文件名}.json")
                    if os.path.exists(old_json_path):
                        os.rename(old_json_path, new_json_path)
                except Exception as e:
                    print(e)
                
                row[index_图片路径 - 1].value = new_image_path
                row[index_文件夹名-1].value = 新文件夹
                row[index_新文件夹-1].value = ''
    
    # 保存修改后的表格
    wb.save(excel_path)


def update_model_json(folder_path):
    excel_name = "model_info.xlsx"
    excel_path = os.path.join(folder_path, excel_name)
    wb = load_workbook(excel_path)
    
    json_name = "model_info.json"
    base_json_path = os.path.join(folder_path, json_name)
    
    # 读取现有的 model_info.json 数据
    with open(base_json_path, 'r', encoding='utf-8') as f:
        base_model_info_json = json.load(f)
    
    # 遍历工作表
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # 获取表头索引映射
        headers = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        header_index = {header: idx - 1 for idx, header in enumerate(headers, start=1)}
        
        for row in ws.iter_rows(min_row=2, values_only=False):
            # 获取文件名及文件夹信息
            file_name = row[0].value  # 文件名
            if not bool(file_name) :
                continue
            
            #! 赋值表达式
            #index = header_index.get('url')
            #value = row[index].value if index is not None and index < len(row) else ''
            #model_entry['url'] = value or ''
            # 等同于
            # value = row[header_index.get('url')] if header_index.get('url') is not None and header_index.get('url') < len(row) else ''
            # 等同于 赋值表达式 walrus operator :=，它允许你在条件表达式内部进行赋值。这种方式可以在一行中完成所有操作，但是可能会降低代码的可读性
            # model_entry['url'] = (row[index].value if (index := header_index.get('url')) is not None and index < len(row) else '') or ''
            
            file_name = (row[index].value if (index := header_index.get('文件名')) is not None and index < len(row) else '') or ''  # 文件名
            file_ext = (row[index].value if (index := header_index.get('拓展名')) is not None and index < len(row) else '') or ''  # 拓展名
            original_name = (row[index].value if (index := header_index.get('原名')) is not None and index < len(row) else '') or '' # 原名
            model_folder = (row[index].value if (index := header_index.get('文件夹名')) is not None and index < len(row) else '') or ''  # 文件夹名
            类型 = (row[index].value if (index := header_index.get('类型')) is not None and index < len(row) else '') or ''  # 类型
            风格 = (row[index].value if (index := header_index.get('风格')) is not None and index < len(row) else '') or ''  # 风格
            用途 = (row[index].value if (index := header_index.get('用途')) is not None and index < len(row) else '') or ''  # 用途
            版本 = (row[index].value if (index := header_index.get('版本')) is not None and index < len(row) else '') or ''  # 版本
            #url = (row[index].value if (index := header_index.get('url')) is not None and index < len(row) else '') or ''  # url
            index_url = header_index.get('url')
            if index_url is not None and index < len(row):
                if row[index_url].hyperlink:
                    url = row[index_url].hyperlink.target
                else:
                    url = row[index_url].value
            图片路径 = (row[index].value if (index := header_index.get('图片路径')) is not None and index < len(row) else '') or ''  # 图片路径
            图片预览 = (row[index].value if (index := header_index.get('图片预览')) is not None and index < len(row) else '') or ''  # 图片预览
            描述 = (row[index].value if (index := header_index.get('描述')) is not None and index < len(row) else '') or ''  # 描述
            sd_link = (row[index].value if (index := header_index.get('SD Link')) is not None and index < len(row) else '') or ''  # SD Link
            特指词 = (row[index].value if (index := header_index.get('特指词')) is not None and index < len(row) else '') or ''  # 特指词
            主描述词 = (row[index].value if (index := header_index.get('主描述词')) is not None and index < len(row) else '') or ''  # 主描述词
            触发词 = (row[index].value if (index := header_index.get('触发词')) is not None and index < len(row) else '') or ''  # 触发词
            可选形象 = (row[index].value if (index := header_index.get('可选形象')) is not None and index < len(row) else '') or ''  # 可选组合
            可选服装 = (row[index].value if (index := header_index.get('可选服装')) is not None and index < len(row) else '') or ''  # 可选服装
            notes = (row[index].value if (index := header_index.get('notes')) is not None and index < len(row) else '') or ''  # notes
            默认权重 = (row[index].value if (index := header_index.get('默认权重')) is not None and index < len(row) else '') or ''  # 默认权重
            权重范围 = (row[index].value if (index := header_index.get('权重范围')) is not None and index < len(row) else '') or ''  # 权重范围
            否定提示词 = (row[index].value if (index := header_index.get('否定提示词')) is not None and index < len(row) else '') or ''  # 否定提示词
            hash_value = (row[index].value if (index := header_index.get('hash')) is not None and index < len(row) else '') or ''  # hash
            喜爱 = (row[index].value if (index := header_index.get('喜爱')) is not None and index < len(row) else '') or ''  # 喜爱
            修改时间 = (row[index].value if (index := header_index.get('修改时间')) is not None and index < len(row) else '') or ''  # 修改时间
            
            # 计算模型的文件夹路径
            model_dir = os.path.join(folder_path, sheet, model_folder) if model_folder else os.path.join(folder_path, sheet)
            # 模型对应的json路径
            model_json_path = os.path.join(model_dir, f'{file_name}.json')
            # 创建 model_info.json 中的键
            model_key = os.path.join(sheet, model_folder, f'{file_name}{file_ext}')
            
            # 创建模型的 JSON 数据（从工作表中提取）
            model_json_data = {
                "pname": original_name,
                "type": 类型,
                "sd version": 类型,
                "风格":风格,
                "用途":用途,
                "版本":版本,
                "url": url,
                "description": 描述,
                "SD Link": sd_link,
                "specific_words": 特指词,
                "main_words": 主描述词,
                "trigger_words": 触发词,
                "activation text": 触发词,
                "可选形象":可选形象,
                "可选服装":可选服装,
                "notes": notes,
                "preferred weight": 默认权重,
                "权重范围":权重范围,
                "negative text": 否定提示词,
                "hash": hash_value,
                "is_favorite": 喜爱,
                "last_modified": 修改时间,
            }
            # model_json_data.update(existing_json_data) 
            # 如果 existing_json_data 中的键在 model_json_data 中不存在，那么会直接将该键值对添加到 model_json_data 中
            # 如果 existing_json_data 中的键在 model_json_data 中已经存在，那么会用 existing_json_data 中对应键的值覆盖 model_json_data 中原来的值
            # ==========================
            # 如果 existing_json_data 中的键在 model_json_data 中已经存在，不进行覆盖model_json_data；
            # 如果 existing_json_data 中的键在 model_json_data 中不存在，直接将该键值对添加到 model_json_data 中
            # 这份代码是将模型同名的json内容追加到要写入model_json文件中的数据
            if os.path.exists(model_json_path):
                with open(model_json_path, 'r', encoding='utf-8') as f:
                    existing_json_data = json.load(f)
                    for key, value in existing_json_data.items():
                        if key not in model_json_data:
                            model_json_data[key] = value
            
            # 更新 model_info.json 数据
            base_model_info_json[model_key] = model_json_data
            #single_base_model_info_json = {model_key:model_json_data}
            #base_model_info_json.update(single_base_model_info_json)
            
            """
            {
                "description": "=== 从Civitai抓取的描述 ===\nI've trained two sets of clothing. The trigger words are as follows\n\ntype A: pantyhose, sweater, jacket\n\ntype B: shirt, skirt, fishnet thighhighs\n\n\nV3.0\n\nThe previous 1.0 version of Marie was really too ugly, so I collected a large amount of new material and made the latest XL version, which took a lot of time... I hope everyone can provide more image feedback to build the SDXL community ecosystem and let more people know the advantages of XL...\n\n\n现在网上的ai玛丽实在是太丑了不忍直视，因此希望大家用这个模型创造出更多可爱的还原的玛丽，更多的优质返图就是对我最大的支持，也希望和更多的朋友讨论如何生成更好看的玛丽。可以关注我的p站和X账号，以后会有更多的图片创作以及优质模型放出...",
                "sd version": "Flux",
                "activation text": "marie rose",
                "preferred weight": 0.76,
                "negative text": "Negative prompt\n",
                "notes": "Notes"
            }
            """
            single_model_json_data = {
                "description": 描述,
                "sd version": 类型,
                "activation text": 触发词,
                "preferred weight": 默认权重,
                "negative text": 否定提示词,
                "notes": notes,
            }
            # 不由代码写入，用月光宝盒的批处理
            # 如果文件不存在，先确保文件夹存在
            #if not os.path.exists(model_json_path):
            #    os.makedirs(os.path.dirname(model_json_path), exist_ok=True)
            #with open(model_json_path, 'w', encoding='utf-8') as f:
            #    json.dump(single_model_json_data, f, ensure_ascii=False, indent=4)
            
            """if os.path.exists(model_json_path):
                with open(model_json_path, 'r+', encoding='utf-8') as f:
                    model_data = json.load(f)
                    model_data.update(model_json_data)  # 使用了model_json_data之前合并数据
                    f.seek(0) # 将文件指针移回文件开头
                    json.dump(model_data, f, ensure_ascii=False, indent=4) # 写入更新后的 JSON 数据
            else:
                # 如果文件不存在，则创建并写入
                os.makedirs(os.path.dirname(model_json_path), exist_ok=True)
                with open(model_json_path, 'w', encoding='utf-8') as f:
                    json.dump(model_json_data, f, ensure_ascii=False, indent=4)"""
    
    # 保存更新后的 model_info.json 文件
    with open(base_json_path, 'w', encoding='utf-8') as f:
        json.dump(base_model_info_json, f, ensure_ascii=False, indent=4)

def format_excel(folder_path):
    excel_name = "model_info.xlsx"
    excel_path = os.path.join(folder_path, excel_name)
    wb = load_workbook(excel_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total_rows = len(rows)  
        columns = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        index_文件名 = columns.index('文件名')+1
        index_拓展名 = columns.index('拓展名')+1
        index_原名 = columns.index('原名')+1
        index_文件夹名 = columns.index('文件夹名')+1
        index_新文件夹 = columns.index('新文件夹')+1
        index_类型 = columns.index('类型')+1
        index_风格 = columns.index('风格')+1
        index_用途 = columns.index('用途')+1
        index_版本 = columns.index('版本')+1
        index_url = columns.index('url')+1
        index_图片路径 = columns.index('图片路径')+1
        index_图片预览 = columns.index('图片预览')+1
        index_描述 = columns.index('描述')+1
        index_SD_Link = columns.index('SD Link')+1
        index_特指词 = columns.index('特指词')+1
        index_主描述词 = columns.index('主描述词')+1
        index_触发词 = columns.index('触发词')+1
        index_可选形象 = columns.index('可选形象')+1
        index_可选服装 = columns.index('可选服装')+1
        index_notes = columns.index('notes')+1
        index_默认权重 = columns.index('默认权重')+1
        index_权重范围 = columns.index('权重范围')+1
        index_否定提示词 = columns.index('否定提示词')+1
        index_hash = columns.index('hash')+1
        index_喜爱 = columns.index('喜爱')+1
        index_修改时间 = columns.index('修改时间')+1
        
        # 让每一个模型默认有一个图片路径
        for row in ws.iter_rows(min_row=2):
            文件名 = row[index_文件名 - 1].value
            文件夹名 = row[index_文件夹名 - 1].value
            # 原始文件夹路径
            if 文件夹名:
                folder = os.path.join(folder_path, sheet, 文件夹名)
            else:
                folder = os.path.join(folder_path, sheet)
            image_path = os.path.join(folder, f"{文件名}.png")
            row[index_图片路径-1].value = image_path
            
        # 设置冻结窗格，固定第一行
        ws.freeze_panes = ws["A2"]
        ws.row_dimensions[1].height = 14
        for row in range(2, total_rows + 1):
            ws.row_dimensions[row].height = 100
        ws.column_dimensions[get_column_letter(index_文件名)].width = 5
        ws.column_dimensions[get_column_letter(index_拓展名)].width = 5
        ws.column_dimensions[get_column_letter(index_原名)].width = 5
        ws.column_dimensions[get_column_letter(index_文件夹名)].width = 5
        ws.column_dimensions[get_column_letter(index_新文件夹)].width = 5
        ws.column_dimensions[get_column_letter(index_类型)].width = 5
        ws.column_dimensions[get_column_letter(index_风格)].width = 5
        ws.column_dimensions[get_column_letter(index_用途)].width = 5
        ws.column_dimensions[get_column_letter(index_版本)].width = 5
        ws.column_dimensions[get_column_letter(index_url)].width = 3
        ws.column_dimensions[get_column_letter(index_图片路径)].width = 3
        ws.column_dimensions[get_column_letter(index_图片预览)].width = 17
        ws.column_dimensions[get_column_letter(index_描述)].width = 40
        ws.column_dimensions[get_column_letter(index_SD_Link)].width = 8
        ws.column_dimensions[get_column_letter(index_特指词)].width = 8
        ws.column_dimensions[get_column_letter(index_主描述词)].width = 8
        ws.column_dimensions[get_column_letter(index_触发词)].width = 8
        ws.column_dimensions[get_column_letter(index_可选形象)].width = 30
        ws.column_dimensions[get_column_letter(index_可选服装)].width = 30
        ws.column_dimensions[get_column_letter(index_notes)].width = 25
        ws.column_dimensions[get_column_letter(index_默认权重)].width = 4
        ws.column_dimensions[get_column_letter(index_权重范围)].width = 10
        ws.column_dimensions[get_column_letter(index_否定提示词)].width = 6
        ws.column_dimensions[get_column_letter(index_hash)].width = 5
        ws.column_dimensions[get_column_letter(index_喜爱)].width = 6
        ws.column_dimensions[get_column_letter(index_修改时间)].width = 14
        # 隐藏列
        ws.column_dimensions[get_column_letter(index_拓展名)].hidden = True
        ws.column_dimensions[get_column_letter(index_图片路径)].hidden = True
        ws.column_dimensions[get_column_letter(index_hash)].hidden = True
        ws.column_dimensions[get_column_letter(index_修改时间)].hidden = True
        for row in ws.iter_rows(min_row=2):
            url_cell = row[index_url-1]
            if url_cell.hyperlink is None:
                url = url_cell.value
                if url:  # 确保单元格值不为空才设置超链接
                    # Hyperlink 对象方式: 灵活但稍复杂，适用于批量设置或需要操作高级功能（如区域超链接）。
                    #hyperlink = Hyperlink(ref=url_cell.coordinate, target=url, display='🌐')
                    #url_cell.hyperlink = hyperlink
                    #url_cell.value = '🌐'
                    
                    # 直接设置方式 (cell.hyperlink): 简洁易用，推荐用于单个或少量超链接操作
                    url_cell.hyperlink = url  # 直接设置单元格的超链接
                    url_cell.style = "Hyperlink"  # 设置超链接样式
                    url_cell.value = '🌐🌐🌐'
        
        for row in ws.iter_rows(min_row=2):
            触发词_cell = row[index_触发词-1]
            特指词_value = row[index_特指词-1].value
            主描述词_value = row[index_主描述词-1].value
            if bool(特指词_value) and bool(主描述词_value):
                触发词_cell.value = ','.join([特指词_value,主描述词_value])
        
        # 设置单元格自动换行
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
    
    sheet_names = ['Stable-diffusion','VAE']
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        ws.column_dimensions[get_column_letter(index_SD_Link)].hidden = True
        ws.column_dimensions[get_column_letter(index_特指词)].hidden = True
        ws.column_dimensions[get_column_letter(index_主描述词)].hidden = True
        ws.column_dimensions[get_column_letter(index_触发词)].hidden = True
        ws.column_dimensions[get_column_letter(index_可选形象)].hidden = True
        ws.column_dimensions[get_column_letter(index_可选服装)].hidden = True
    
    wb.save(excel_path)

def reinsert_image(folder_path):
    excel_name = "model_info.xlsx"
    excel_path = os.path.join(folder_path, excel_name)
    wb = load_workbook(excel_path)
    
    sheets = wb.sheetnames
    # 先遍历所有工作表清空已有图片
    for sheet in sheets:
        ws = wb[sheet]
        # 获取图片的列表
        images = list(ws._images)  # 创建一个副本以避免在迭代时修改列表长度
        for drawing in images:
            ws._images.remove(drawing)
    """# 先遍历所有工作表清空已有图片，该方法会有概率会遗漏删除
    # 可能是因为在遍历 _images 列表并删除元素时，列表的长度发生了变化，导致索引错乱，使得部分元素没有被遍历到
    for sheet in sheets:
        ws = wb[sheet]
        for drawing in ws._images:
            ws._images.remove(drawing)"""
            
    """# 创建新的工作簿
    # 太复杂，没有采用，代码未完成，仅参考
    new_wb = Workbook()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        new_ws = new_wb.create_sheet(sheet_name)
        # 复制所有数据和样式
        for row in ws.iter_rows():
            for cell in row:
                new_cell = new_ws[cell.coordinate]
                # 复制值
                new_cell.value = cell.value
                # 复制样式
                if cell.has_style:
                # 字体
                new_cell.font = cell.font
                # 边框
                new_cell.border = cell.border
                # 填充
                new_cell.fill = cell.fill
                # 对齐方式
                new_cell.alignment = cell.alignment
                # 数字格式
                new_cell.number_format = cell.number_format
                # 保护属性
                new_cell.protection = cell.protection
        # 复制列宽
        for col_letter, col_dim in ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
        # 复制行高
        for row_num, row_dim in ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = row_dim.height
        new_wb = Workbook()
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            new_ws = new_wb.create_sheet(sheet_name)
            # 复制所有数据
            for row in ws.iter_rows():
                for cell in row:
                    new_ws[cell.coordinate].value = cell.value
    # 删除默认生成的工作表（通常名为'Sheet'）
    default_sheet_name = 'Sheet'
    for sheet in wb.sheetnames:
        if default_sheet_name in sheet:
            default_sheet = wb[sheet]
            wb.remove(default_sheet)"""
    
    
    for sheet in sheets:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, max_col=12):
            if row[0].value is None :
                break
            for col_index, cell in enumerate(row, start=1):
                value = cell.value
                if value is None or not isinstance(value, str):
                    continue
                if value.startswith('E:\models') and 'png' in value and os.path.exists(value):
                    img = OpenpyxlImage(value)
                    # 调整单元格宽高
                    img_width = img.width
                    img_height = img.height
                    img_scale = img_width / img_height
                    # 固定高度和高度都为 100 px
                    height_pt = 100  # 目标高度为 100 px
                    height_px = 100 * (4 / 3)
                    width_ch = height_px / 8
                    
                    # 单元格限制：高度 133 px，宽度 133 px
                    # 等同于：行高 100 pt，列宽 16.625 字符
                    max_height_px = height_px
                    max_width_px = height_px
                    # 根据比例调整宽度和高度
                    if img_width > max_width_px or img_height > max_height_px:
                        # 判断以哪个方向缩放
                        if img_width / max_width_px > img_height / max_height_px:
                            # 按宽度缩放
                            new_width = max_width_px
                            new_height = int(new_width / img_scale)
                        else:
                            # 按高度缩放
                            new_height = max_height_px
                            new_width = int(new_height * img_scale)
                    else:
                        # 图片小于单元格，保持原始大小
                        new_width = img_width
                        new_height = img_height
                    
                    # 设置图片的新宽度和高度
                    img.width = new_width
                    img.height = new_height
                    
                    colname = get_column_letter(col_index+1)
                    rowindex = cell.row
                    ws.column_dimensions[colname].width = width_ch  # 列宽单位为字符，像素单位需除以 8
                    ws.row_dimensions[rowindex].height = height_pt  # 行高单位为磅，像素单位需除以 4/3
                    ws.add_image(img, f"{colname}{rowindex}")
                    print(f'插入图片: {value} -> 单元格 {colname}{rowindex}')
        
        for row in ws.iter_rows(min_row=2, min_col=13):
            if row[0].value is None :
                break
            for col_index, cell in enumerate(row, start=13):
                value = cell.value
                if value is None or not isinstance(value, str):
                    continue
                if value.startswith('D:\AI Tech') and 'png' in value and os.path.exists(value):
                    resized_img_path = value.replace(".png", "_compressed.png")
                    # 检查压缩图片路径是否存在
                    if not os.path.exists(resized_img_path):
                        try:
                            # 压缩图片
                            max_width = 768
                            max_height = 768
                            # 打开原始图片
                            img = PILImage.open(value)
                            # 获取原始图片的宽高
                            width, height = img.size
                            # 计算缩放比例，保持宽高比进行缩放，使图片最长边不超过指定值
                            ratio = min(max_width / width if width > max_height else 1,
                                        max_height / height if height > max_width else 1)
                            new_width = int(width * ratio)
                            new_height = int(height * ratio)
                            # 进行缩放
                            resized_img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
                            # 保存压缩后的图片
                            resized_img.save(resized_img_path)
                        except Exception as e:
                            print(f"处理图片 {value} 时出错: {e}")
                    
                    img = OpenpyxlImage(resized_img_path)
                    # 调整单元格宽高
                    img_width = img.width
                    img_height = img.height
                    img_scale = img_width / img_height
                    # 固定高度和高度都为 100 px
                    height_pt = 100  # 目标高度为 100 px
                    height_px = 100 * (4 / 3)
                    width_ch = height_px / 8
                    
                    # 单元格限制：高度 133 px，宽度 133 px
                    # 等同于：行高 100 pt，列宽 16.625 字符
                    max_height_px = height_px
                    max_width_px = height_px
                    # 根据比例调整宽度和高度
                    if img_width > max_width_px or img_height > max_height_px:
                        # 判断以哪个方向缩放
                        if img_width / max_width_px > img_height / max_height_px:
                            # 按宽度缩放
                            new_width = max_width_px
                            new_height = int(new_width / img_scale)
                        else:
                            # 按高度缩放
                            new_height = max_height_px
                            new_width = int(new_height * img_scale)
                    else:
                        # 图片小于单元格，保持原始大小
                        new_width = img_width
                        new_height = img_height
                    
                    # 设置图片的新宽度和高度
                    img.width = new_width
                    img.height = new_height
                    
                    ws.column_dimensions[get_column_letter(col_index)].width = 3
                    colname = get_column_letter(col_index+1)
                    rowindex = cell.row
                    ws.column_dimensions[colname].width = width_ch  # 列宽单位为字符，像素单位需除以 8
                    ws.row_dimensions[rowindex].height = height_pt  # 行高单位为磅，像素单位需除以 4/3
                    ws.add_image(img, f"{colname}{rowindex}")
                    print(f'插入图片: {value} -> 单元格 {colname}{rowindex}')
    
    wb.save(excel_path)

folder_path = "E:\models"
# JSON写入表格-OK
json_to_execl(folder_path)

# 文件重命名-OK
rename_filenames(folder_path)

# 移动文件-OK
move_to_newfolder(folder_path)

# 更新JSON文件-OK
update_model_json(folder_path)

# 格式化表格样式-OK
format_excel(folder_path)

# 重新插入图片-OK
reinsert_image(folder_path)